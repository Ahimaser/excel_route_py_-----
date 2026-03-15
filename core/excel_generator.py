"""
excel_generator.py — Генератор XLS файлов маршрутов.

Оглавление секций:
  1. Кэш стилей, утилиты (пути, даты, calc_pcs)
  2. Парсинг адресов (extract_house_number, _extract_house_parts)
  3. Генерация общих маршрутов (generate_general_routes)
  4. Этикетки (labels_preview, labels_preview_rows, generate_labels_from_templates)
  5. Файлы по отделам (_write_dept_sheet, generate_dept_files, generate_single_dept_file)

Форматы файлов по отделам:
  "wide"  (Формат 1) — каждый продукт в отдельном столбце
  "rows"  (Формат 2) — строчный (маршрут + продукты)

Сортировка: по убыванию номера маршрута (числовая).
"""
from __future__ import annotations

import copy
import json
import logging
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, timedelta, datetime
import time
from typing import Any

import xlwt

from core import data_store

log = logging.getLogger("excel_generator")

ROUTE_SIGN = "\u2116"

# ─────────────────────────── Кэш стилей ──────────────────────────────────

_STYLES: dict[str, xlwt.XFStyle] | None = None
_STYLES_FONT_PT: int | None = None


def _get_styles() -> dict[str, xlwt.XFStyle]:
    """Возвращает набор стилей. Размер шрифта из настроек (по умолчанию 12pt). Кэш по font_pt."""
    global _STYLES, _STYLES_FONT_PT
    font_pt = 12
    try:
        v = data_store.get_setting("defaultFontSize")
        if v is not None:
            font_pt = max(8, min(24, int(v)))
    except Exception:
        pass
    if _STYLES is not None and _STYLES_FONT_PT == font_pt:
        return _STYLES
    height = font_pt * 20  # xlwt: height в 1/20 пункта

    font_bold = xlwt.Font()
    font_bold.bold = True
    font_bold.height = height

    font_normal = xlwt.Font()
    font_normal.height = height

    align_wrap = xlwt.Alignment()
    align_wrap.wrap = xlwt.Alignment.WRAP_AT_RIGHT
    align_wrap.vert = xlwt.Alignment.VERT_TOP

    align_center = xlwt.Alignment()
    align_center.horz = xlwt.Alignment.HORZ_CENTER
    align_center.vert = xlwt.Alignment.VERT_TOP

    align_top = xlwt.Alignment()
    align_top.vert = xlwt.Alignment.VERT_TOP

    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN

    def _make(font, alignment, has_borders=True):
        s = xlwt.XFStyle()
        s.font = font
        s.alignment = alignment
        if has_borders:
            s.borders = borders
        return s

    # Жёлтый фон для номера маршрута в этикетках
    pattern_yellow = xlwt.Pattern()
    pattern_yellow.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern_yellow.pattern_fore_colour = xlwt.Style.colour_map["yellow"]
    style_yellow = xlwt.XFStyle()
    style_yellow.font = font_normal
    style_yellow.alignment = align_top
    style_yellow.borders = borders
    style_yellow.pattern = pattern_yellow

    # Хвостовая строка в общих маршрутах: № + номер + дом/строение/корпус (жирный 30pt, по центру)
    font_tail = xlwt.Font()
    font_tail.bold = True
    font_tail.height = 30 * 20  # 30pt
    align_center_tail = xlwt.Alignment()
    align_center_tail.horz = xlwt.Alignment.HORZ_CENTER
    align_center_tail.vert = xlwt.Alignment.VERT_TOP
    style_tail = xlwt.XFStyle()
    style_tail.font = font_tail
    style_tail.alignment = align_center_tail
    style_tail.borders = borders

    # num_bold — для ячеек «количество + Шт» (продукт колонка на каждый)
    style_num_bold = xlwt.XFStyle()
    style_num_bold.font = font_bold
    style_num_bold.alignment = align_center
    style_num_bold.borders = borders

    _STYLES = {
        "header":      _make(font_bold,   align_center),
        "header_wrap": _make(font_bold,   align_wrap),
        "cell":        _make(font_normal, align_top),
        "cell_wrap":   _make(font_normal, align_wrap),
        "num":         _make(font_normal, align_center),  # количество и Шт — по центру ячейки
        "num_bold":    style_num_bold,   # количество + Шт жирным (productQty с pcs)
        "title":       _make(font_bold,   align_top, has_borders=False),
        "cell_yellow": style_yellow,
        "tail_line":   style_tail,
    }
    _STYLES_FONT_PT = font_pt
    return _STYLES


def _apply_page_margins(ws: xlwt.Worksheet, for_labels: bool = False) -> None:
    """Применяет отступы страницы из настроек (см). Не применяется к этикеткам."""
    if for_labels:
        return
    try:
        top = float(data_store.get_setting("defaultMarginTop") or 1.5)
        left = float(data_store.get_setting("defaultMarginLeft") or 1.5)
        bottom = float(data_store.get_setting("defaultMarginBottom") or 0.5)
        right = float(data_store.get_setting("defaultMarginRight") or 0.5)
    except (TypeError, ValueError):
        top, left, bottom, right = 1.5, 1.5, 0.5, 0.5
    inch = 1.0 / 2.54
    ws.set_top_margin(top * inch)
    ws.set_left_margin(left * inch)
    ws.set_bottom_margin(bottom * inch)
    ws.set_right_margin(right * inch)


# ─────────────────────────── Утилиты ──────────────────────────────────────

def _tomorrow() -> date:
    return date.today() + timedelta(days=1)


def _format_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def _type_label(file_type: str) -> str:
    return "Увеличение (Довоз)" if file_type == "increase" else "Основной"


def _type_suffix(file_type: str) -> str:
    return "УВ" if file_type == "increase" else "ОСН"


def _dept_special_mode_raw(name: str, explicit_mode: str | None = None) -> str:
    mode = (explicit_mode or "").strip().lower()
    if mode:
        if mode in ("chistchenka", "sypuchka", "polufabricates", "polufabrikaty"):
            return "polufabricates" if mode in ("polufabricates", "polufabrikaty") else mode
        return mode
    lname = (name or "").strip().lower()
    if "чищен" in lname:
        return "chistchenka"
    if "полуфаб" in lname:
        return "polufabricates"
    if "сыпуч" in lname:
        return "sypuchka"
    return "default"


def _build_dept_mode_map() -> dict[str, str]:
    mode_map: dict[str, str] = {}
    for dept in data_store.get_ref("departments") or []:
        dkey = str(dept.get("key") or "")
        if dkey:
            mode_map[dkey] = _dept_special_mode_raw(
                str(dept.get("name") or ""),
                str(dept.get("labelPrintMode") or ""),
            )
        for sub in dept.get("subdepts", []):
            skey = str(sub.get("key") or "")
            if skey:
                mode_map[skey] = _dept_special_mode_raw(
                    str(sub.get("name") or ""),
                    str(sub.get("labelPrintMode") or ""),
                )
    return mode_map


def get_dept_special_mode(dept_key: str | None) -> str:
    """Возвращает спец-режим отдела: default/chistchenka/sypuchka/polufabricates."""
    if not dept_key:
        return "default"
    return _build_dept_mode_map().get(str(dept_key), "default")


def get_routes_date_str() -> str:
    """Дата маршрутов (как в файлах): обычно следующий день."""
    return _format_date(_tomorrow())


def get_routes_day_folder(base_dir: str, date_str: str | None = None) -> str:
    """Папка дня: base_dir/Маршруты {date}. Не дублирует, если base_dir уже заканчивается на неё."""
    date_part = date_str or get_routes_date_str()
    expected_tail = f"Маршруты {date_part}"
    if os.path.basename(os.path.normpath(base_dir)) == expected_tail:
        return base_dir
    return os.path.join(base_dir, expected_tail)


def get_routes_type_folder(base_dir: str, file_type: str, date_str: str | None = None) -> str:
    """Папка типа: day_dir/Основные или day_dir/Увеличение. Не дублирует, если base_dir уже она."""
    sub = "Увеличение" if file_type == "increase" else "Основные"
    if os.path.basename(os.path.normpath(base_dir)) == sub:
        return base_dir
    day_dir = get_routes_day_folder(base_dir, date_str)
    return os.path.join(day_dir, sub)


def get_general_routes_path(base_dir: str, file_type: str, date_str: str | None = None) -> str:
    date_part = date_str or get_routes_date_str()
    type_dir = get_routes_type_folder(base_dir, file_type, date_part)
    return os.path.join(type_dir, f"Общие маршруты {date_part}.xls")


def get_dept_routes_path(
    base_dir: str,
    file_type: str,
    dept_name: str,
    date_str: str | None = None,
    parent_dept_name: str | None = None,
) -> str:
    """
    Путь к файлу маршрутов отдела.
    Файлы по отделам — в той же папке Основные/Увеличение, что и общие маршруты.
    Для отдела: Основные/{отдел}/Сборка {отдел} {дата}.xls
    Для подотдела: Основные/{родительский_отдел}/Сборка {подотдел} {дата}.xls (файл в папке отдела)
    """
    date_part = date_str or get_routes_date_str()
    type_dir = get_routes_type_folder(base_dir, file_type, date_str)
    safe_name = _safe_filename(dept_name)
    folder_name = _safe_filename(parent_dept_name) if parent_dept_name else safe_name
    dept_folder = os.path.join(type_dir, folder_name)
    return os.path.join(dept_folder, f"Сборка {safe_name} {date_part}.xls")


def get_dept_product_file_path(
    base_dir: str,
    file_type: str,
    dept_name: str,
    product_names: list[str],
    date_str: str | None = None,
    parent_dept_name: str | None = None,
) -> str:
    """
    Путь к файлу маршрутов по группе продуктов (режим «разделить по продуктам»).
    Файлы в папке отдела: Основные/{отдел}/Сборка {продукт1}, {продукт2} {дата}.xls
    Для подотдела: Основные/{родитель}/Сборка {подотдел} — {продукт1}, {продукт2} {дата}.xls
    """
    date_part = date_str or get_routes_date_str()
    type_dir = get_routes_type_folder(base_dir, file_type, date_str)
    products_part = ", ".join(product_names) if product_names else "Продукт"
    safe_products = _safe_filename(products_part)
    if len(safe_products) > 80:
        safe_products = safe_products[:77] + "..."
    folder_name = _safe_filename(parent_dept_name) if parent_dept_name else _safe_filename(dept_name)
    dept_folder = os.path.join(type_dir, folder_name)
    if parent_dept_name:
        safe_dept = _safe_filename(dept_name)
        filename = f"Сборка {safe_dept} — {safe_products} {date_part}.xls"
    else:
        filename = f"Сборка {safe_products} {date_part}.xls"
    return os.path.join(dept_folder, filename)


def calc_pcs(quantity: float, pcs_per_unit: float, round_up: bool = True) -> int:
    """
    Рассчитывает количество штук (старая логика: ceil/floor по всему отношению).
    """
    if pcs_per_unit <= 0 or quantity <= 0:
        return 0
    ratio = quantity / pcs_per_unit
    if round_up:
        return max(0, int(math.ceil(ratio)))
    return max(0, int(math.floor(ratio)))


def calc_pcs_tail(
    quantity: float,
    pcs_per_unit: float,
    round_tail_from: float,
) -> int:
    """
    Рассчитывает штуки по логике «хвостик ШК/СД».

    Хвостик — граница остатка (в тех же единицах, что количество: кг, л и т.д.).
    При остатке от этой границы включительно округление в большую сторону (+1 шт).
    Пример: 1 шт = 0.7 кг, хвостик ШК = 0.35 → при остатке ≥ 0.35 кг добавляем 1 шт.

    Целые штуки = floor(quantity / pcs_per_unit).
    Остаток = quantity - целые * pcs_per_unit.
    round_tail_from = 0: при любом ненулевом остатке +1 шт (как ceil).
    round_tail_from > 0: при остатке >= round_tail_from +1 шт.
    """
    if pcs_per_unit <= 0 or quantity <= 0:
        return 0
    full = int(math.floor(quantity / pcs_per_unit))
    remainder = quantity - full * pcs_per_unit
    if round_tail_from <= 0:
        add_one = remainder > 0
    else:
        add_one = remainder >= round_tail_from
    if add_one:
        return full + 1
    return full


def _apply_pcs(routes: list[dict], prod_map: dict[str, dict], group_dept_key: str | None = None) -> list[dict]:
    """
    Добавляет к продуктам маршрутов displayQuantity (с учётом множителя замены) и pcs.
    prod_map: {name: product_settings_dict}. Коэффициент замены (quantityMultiplier), напр. 1.25
    для пересчёта очищенных → грязные: отображаемое количество = количество × коэффициент.
    Округление берётся по категории маршрута (ШК/СД), а для некоторых учреждений —
    в большую сторону по проценту отдела (см. is_always_round_up_institution, get_institution_round_percent).
    group_dept_key: ключ отдела/подотдела при генерации по группе — используется как fallback для режима (напр. полуфабрикаты).
    """
    dept_mode_map = _build_dept_mode_map()
    aliases = data_store.get_aliases()
    for route in routes:
        route_cat = route.get("routeCategory") or "ШК"
        addr = route.get("address", "")
        force_round_up = is_always_round_up_institution(addr)
        for prod in route.get("products", []):
            canonical = aliases.get(prod["name"], prod["name"])
            sp = prod_map.get(canonical, prod_map.get(prod["name"], {}))
            dept_key = sp.get("deptKey") or group_dept_key
            dept_mode = dept_mode_map.get(str(dept_key or ""), "default")
            if dept_mode == "default" and data_store.is_subdept_polufabricates(dept_key):
                dept_mode = "polufabricates"
            qty = prod.get("quantity")
            mult = float(sp.get("quantityMultiplier", 1.0) or 1.0)
            if qty is not None:
                try:
                    display_qty = float(qty) * mult
                    prod["displayQuantity"] = display_qty
                except (ValueError, TypeError):
                    prod["displayQuantity"] = qty
            else:
                prod["displayQuantity"] = qty

            pcs = None
            pcs_tail = None
            if sp.get("showPcs") and prod.get("unit", "").lower() != "шт":
                eff_qty = prod.get("displayQuantity")
                if eff_qty is not None:
                    try:
                        val = float(eff_qty)
                        pcu = float(sp.get("pcsPerUnit", 1))
                        if pcu <= 0:
                            pcs = 0
                        else:
                            if dept_mode == "polufabricates":
                                # Для полуфабрикатов округления нет: показываем целые шт + хвостик.
                                pcs = max(0, int(math.floor(val / pcu)))
                                pcs_tail = max(0.0, float(val - pcs * pcu))
                            else:
                                # Порог: ниже — 0 шт, от порога и выше — расчёт по округлению
                                min_qty = sp.get("minQtyForPcs")
                                if min_qty is not None and min_qty > 0:
                                    threshold = float(min_qty)
                                else:
                                    unit_lower = (prod.get("unit") or "").strip().lower()
                                    threshold = 0.2 if unit_lower in ("кг", "л", "kg", "l") else 0
                                if val < threshold:
                                    pcs = 0
                                else:
                                    if force_round_up:
                                        # Для заданных учреждений — в большую по % отдела
                                        pct = get_institution_round_percent(dept_key)
                                        round_tail = pcu * (pct / 100.0)
                                        pcs = calc_pcs_tail(val, pcu, round_tail)
                                    else:
                                        round_tail = (
                                            sp.get("roundTailFromСД") if route_cat == "СД" else sp.get("roundTailFromШК")
                                        )
                                        if round_tail is not None:
                                            round_tail = float(round_tail)
                                            pcs = calc_pcs_tail(val, pcu, round_tail)
                                        else:
                                            round_up = (
                                                sp.get("roundUpСД") if "roundUpСД" in sp else sp.get("roundUp", True)
                                                if route_cat == "СД"
                                                else sp.get("roundUpШК") if "roundUpШК" in sp else sp.get("roundUp", True)
                                            )
                                            pcs = calc_pcs(val, pcu, bool(round_up))
                    except (ValueError, TypeError):
                        pass
            prod["pcs"] = pcs
            prod["pcsTail"] = pcs_tail
    return routes


# Символы, недопустимые в имени листа Excel: \ / ? * [ ]
_SHEET_NAME_FORBIDDEN = re.compile(r'[\\/?*\[\]]')


def _safe_save_workbook(wb: xlwt.Workbook, save_path: str) -> None:
    """
    Сохраняет файл с перезаписью: сначала во временный файл, затем replace.
    Обеспечивает обновление при повторном сохранении (в т.ч. если файл открыт в другом процессе).
    """
    abs_path = os.path.abspath(save_path)
    dir_ = os.path.dirname(abs_path)
    fd, tmp_path = tempfile.mkstemp(suffix=".xls", dir=dir_ or None, prefix=".xlwt_")
    try:
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, abs_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _safe_sheet_name(name: str) -> str:
    """Имя листа Excel: макс 31 символ, без \\ / ? * [ ]."""
    s = str(name).strip()
    s = _SHEET_NAME_FORBIDDEN.sub("_", s)
    s = s.strip("_") or "Лист"
    return s[:31]


def _unique_sheet_name(name: str, used: set[str]) -> str:
    """Генерирует уникальное имя листа (макс 31 символ, без недопустимых символов)."""
    base = _safe_sheet_name(name)
    if len(base) > 28:
        base = base[:28]
    candidate = base
    counter = 2
    while candidate in used:
        candidate = f"{base[:24]}_{counter}"
        counter += 1
    used.add(candidate)
    return candidate


def _set_col_width(sheet: xlwt.Worksheet, col: int, width_chars: int) -> None:
    sheet.col(col).width = min(width_chars * 256, 65535)


def _sort_routes(routes: list[dict], sort_asc: bool = True) -> list[dict]:
    """Сортирует маршруты по номеру маршрута (числовая сортировка).
    sort_asc=True (по умолчанию) — по возрастанию.
    sort_asc=False — по убыванию.
    Маршруты с неопределённым номером — всегда в начало.
    """
    def _key(r: dict):
        num = r.get("routeNum", "")
        try:
            n = int(str(num).strip())
            return (1, n if sort_asc else -n)
        except (ValueError, TypeError):
            return (0, 0)
    return sorted(routes, key=_key)


def _safe_filename(name: str) -> str:
    """Убирает запрещённые символы из имени файла."""
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def extract_house_number(address: str) -> str:
    """
    Извлекает из адреса номер дома, строения, корпуса и/или цифры перед № (U+2116).
    Поддерживает смешанный формат: "дом 3 строение 2", "д. 2 корпус 1", "д.5", "стр. 2" и т.д.
    Возвращает одну строку с найденными значениями через запятую (например "3, стр. 2, корп. 1").
    """
    return _extract_house_parts(address, include_route_num=True)


def _strip_route_letter_from_house(house_part: str) -> str:
    """
    Убирает букву М (маршрут) после номера дома.
    д.6А, д.8Б, д. 6 А — оставляем букву. д.21М, д.5Марш — убираем (М = маршрут).
    Между цифрами и буквой может быть пробел (д. 6 А → 6А).
    """
    if not house_part:
        return house_part
    s = house_part.strip()
    m = re.match(r"^(\d+(?:/\d+)?)\s*([а-яА-Яa-zA-Z]*)\s*$", s)
    if not m:
        return house_part
    digits, letters = m.group(1), (m.group(2) or "").strip()
    if not letters:
        return digits
    first = letters[0].upper()
    if first == "М" or first == "M":  # М = маршрут, не включаем
        return digits
    return digits + letters


def _extract_house_parts(address: str, include_route_num: bool = True) -> str:
    """
    Извлекает дом/строение/корпус/владение из адреса.
    После номера дома может быть буква (д.6А, д.8Б стр.1). Буква М (маршрут) не включается.
    Не захватывает «стр»/«корп»/«вл» как букву дома (д.34 стр.1 → 34, стр. 1).
    include_route_num: включать ли цифры перед № (номер маршрута в адресе).
    """
    if not address:
        return ""
    s = str(address).strip()
    parts: list[str] = []

    # дом / д. / д / д4 — дом 3, д. 5, д.5а, д.6А, д. 6 А, д 8Б, д4
    # Буква дома — только А, Б и т.д.; не «стр» (строение), «корп» (корпус), «вл» (владение), «ул» (улица)
    m = re.search(
        r"(?:^|[^\w])(?:дом\s*|д\.\s*|д\s+|д(?=\d))"
        r"(\d+(?:/\d+)?)\s*"
        r"(?![сС]тр\.?|[кК]орп\.?|[вВ]л\.?|[уУ]л\.?)([а-яА-Яa-zA-Z]*)",
        s, re.IGNORECASE
    )
    if m:
        digits, letters = m.group(1), (m.group(2) or "").strip()
        raw = digits + letters
        parts.append(_strip_route_letter_from_house(raw))

    # строение / строен. / стр. — только первое вхождение (без дубликатов)
    m = re.search(r"(?:строение|строен\.?|стр\.)\s*(\d+)", s, re.IGNORECASE)
    if m:
        parts.append(f"стр. {m.group(1)}")

    # корпус / корп. — только первое вхождение (корп. 1, корпус 2)
    m = re.search(r"корп(?:ус)?\.?\s*(\d+)", s, re.IGNORECASE)
    if m:
        parts.append(f"корп. {m.group(1)}")

    # владение / влад. / вл. — влад. 1, владение 2, вл. 3
    m = re.search(r"(?:владение|влад\.?|вл\.)\s*(\d+)", s, re.IGNORECASE)
    if m:
        parts.append(f"влад. {m.group(1)}")

    # цифры перед символом № (U+2116) — только если нужны (для хвостовой строки не включаем, т.к. номер уже есть)
    if include_route_num:
        m = re.search(r"(\d+)\s*[№\u2116]", s)
        if m:
            num = m.group(1)
            if num not in parts:
                parts.append(num)

    return ", ".join(parts) if parts else ""


def is_always_round_up_institution(address: str) -> bool:
    """
    Проверяет, включён ли адрес в список учреждений с округлением шт в большую сторону.
    Ключ учреждения: первые 3–4 цифры (109/1 → 109, 1391/2 → 1391).
    Адреса из excludeRoundUpAddresses исключаются.
    """
    key = data_store.get_institution_key_from_address(address or "")
    if not key:
        return False
    try:
        codes = data_store.get_setting("alwaysRoundUpInstitutions") or []
        if key not in codes:
            return False
        excluded = set(data_store.get_setting("excludeRoundUpAddresses") or [])
        if (address or "").strip() in excluded:
            return False
        return True
    except Exception:
        return False


def get_institution_round_percent(dept_key: str | None) -> float:
    """
    % от 1 шт для округления в учреждениях. Зависит от отдела (roundUpPercentByDept).
    """
    return data_store.get_round_up_percent_for_dept(dept_key)


def _label_print_mode_for_dept(dept_key: str | None, departments_ref: list | None) -> str:
    """Режим печати этикеток: default, chistchenka, sypuchka (по имени или labelPrintMode)."""
    if not dept_key or not departments_ref:
        return "default"
    for dept in departments_ref:
        if dept.get("key") == dept_key:
            mode = dept.get("labelPrintMode")
            if mode in ("chistchenka", "sypuchka"):
                return mode
            name = (dept.get("name") or "").lower()
            if "очищенные" in name:
                return "chistchenka"
            if "сыпучка" in name:
                return "sypuchka"
            return "default"
        for sub in dept.get("subdepts", []):
            if sub.get("key") == dept_key:
                mode = sub.get("labelPrintMode")
                if mode in ("chistchenka", "sypuchka"):
                    return mode
                name = (sub.get("name") or "").lower()
                if "очищенные" in name:
                    return "chistchenka"
                if "сыпучка" in name:
                    return "sypuchka"
                return "default"
    return "default"


def _label_rules_for_dept(dept_key: str | None, departments_ref: list | None) -> dict:
    """
    Возвращает правила этикеток для отдела/подотдела: labelRules или пустой dict.
    labelRules может содержать chistchenka: {maxKgPerLabel} (макс. кг на этикетку; дубликаты по 5 кг, остаток — последняя этикетка),
    sypuchka: {thresholdKg, labelAbove, labelBelow}.
    """
    if not dept_key or not departments_ref:
        return {}
    for dept in departments_ref:
        if dept.get("key") == dept_key:
            return dict(dept.get("labelRules") or {})
        for sub in dept.get("subdepts", []):
            if sub.get("key") == dept_key:
                return dict(sub.get("labelRules") or {})
    return {}


def _dept_display_name(dept_key: str | None, departments_ref: list | None) -> str:
    """Возвращает отображаемое имя отдела/подотдела по key."""
    if not dept_key or not departments_ref:
        return ""
    for dept in departments_ref:
        if dept.get("key") == dept_key:
            return dept.get("name") or dept_key
        for sub in dept.get("subdepts", []):
            if sub.get("key") == dept_key:
                parent = dept.get("name") or ""
                sub_name = sub.get("name") or dept_key
                return f"{parent} / {sub_name}" if parent else sub_name
    return dept_key


def _route_sort_key_labels(route_num: str) -> tuple[int, int, str]:
    """Ключ сортировки номера маршрута по возрастанию (для этикеток: сначала по адресу, затем по этому ключу)."""
    try:
        return (1, int(str(route_num).strip()), str(route_num))
    except (ValueError, TypeError):
        return (0, 0, str(route_num))


def _label_sort_key_route(route: dict) -> tuple[str, int, int, str]:
    """Ключ сортировки для этикеток: сначала адрес, затем номер маршрута по возрастанию."""
    address = (route.get("address") or "").strip()
    route_num = str(route.get("routeNum", ""))
    return (address, *_route_sort_key_labels(route_num))


def _load_template_matrix_xlsx(
    template_path: str,
) -> tuple[int, int, list, int, list[int], list[int], list[tuple[int, int, int, int]], list[int]]:
    """Загрузка шаблона XLSX через openpyxl. Возвращает тот же формат, что и _load_template_matrix."""
    from openpyxl import load_workbook
    wb = load_workbook(template_path, data_only=True)
    sheet = wb.active
    matrix: list[list] = []
    row_heights: list[int] = []
    source_rows: list[int] = []
    last_filled = -1
    orig_to_matrix: dict[int, int] = {}
    ncols = 0
    for r_idx, row in enumerate(sheet.iter_rows()):
        if r_idx >= 65536:
            break
        rd = sheet.row_dimensions.get(r_idx + 1)
        row_hidden = getattr(rd, "hidden", False) if rd else False
        row_height = getattr(rd, "height", 15) if rd else 15
        row_height = row_height or 15
        if row_hidden or row_height <= 0:
            continue
        orig_to_matrix[r_idx] = len(matrix)
        row_vals: list = []
        for c_idx, cell in enumerate(row):
            if cell.value is None:
                row_vals.append("")
            elif isinstance(cell.value, (int, float)):
                v = cell.value
                row_vals.append(int(v) if isinstance(v, float) and v == int(v) else v)
            else:
                row_vals.append(str(cell.value).strip() if cell.value else "")
        ncols = max(ncols, len(row_vals))
        matrix.append(row_vals)
        row_heights.append(int(row_height * 20) if row_height else 0)
        source_rows.append(r_idx)
        if any(v != "" for v in row_vals):
            last_filled = len(matrix) - 1
    nrows_trimmed = (last_filled + 1) if last_filled >= 0 else (1 if matrix else 0)
    matrix = matrix[:nrows_trimmed]
    row_heights = row_heights[:nrows_trimmed]
    source_rows = source_rows[:nrows_trimmed]
    for row in matrix:
        while len(row) < ncols:
            row.append("")
    from openpyxl.utils import get_column_letter
    col_widths = []
    for c in range(ncols):
        letter = get_column_letter(c + 1)
        try:
            cd = sheet.column_dimensions.get(letter)
            w = getattr(cd, "width", None) if cd else None
        except Exception:
            w = None
        col_widths.append(int((w or 10) * 256))
    merges = []
    for mrng in getattr(sheet.merged_cells, "ranges", []) or []:
        min_r, min_c = mrng.min_row - 1, mrng.min_col - 1
        max_r, max_c = mrng.max_row - 1, mrng.max_col - 1
        visible = [orig_to_matrix[r] for r in range(min_r, max_r + 1) if r in orig_to_matrix]
        if not visible:
            continue
        m_r1, m_r2 = visible[0], visible[-1]
        if m_r1 >= nrows_trimmed:
            continue
        m_r2 = min(m_r2, nrows_trimmed - 1)
        merges.append((m_r1, m_r2, min_c, max_c))
    return len(matrix), ncols, matrix, last_filled, row_heights, col_widths, merges, source_rows


def _load_template_matrix(
    template_path: str,
) -> tuple[int, int, list, int, list[int], list[int], list[tuple[int, int, int, int]], list[int]]:
    """
    Загружает шаблон XLS/XLSX в матрицу.
    Скрытые строки (hidden=True или height=0) пропускаются полностью.
    Пустые строки в конце обрезаются.

    Возвращает:
        nrows       — число строк в обрезанной матрице
        ncols       — число столбцов
        matrix      — list[list] значений ячеек
        last_filled — индекс последней непустой строки
        row_heights — высоты строк (1/20 пункта) для каждой видимой строки матрицы
        col_widths  — ширины столбцов (1/256 символа)
        merges      — объединённые ячейки в пространстве матрицы:
                      list of (r1, r2, c1, c2) включительно
        source_rows — исходные номера строк шаблона (0-based) для каждой строки матрицы
    """
    if (template_path or "").lower().endswith(".xlsx"):
        return _load_template_matrix_xlsx(template_path)
    import xlrd
    # formatting_info=True нужен для rowinfo_map, colinfo_map, merged_cells
    wb = xlrd.open_workbook(template_path, formatting_info=True)
    sheet = wb.sheet_by_index(0)
    ncols = sheet.ncols
    matrix: list[list] = []
    row_heights: list[int] = []
    source_rows: list[int] = []
    last_filled = -1

    # Карта: исходный индекс строки файла → индекс в матрице (только видимые строки)
    orig_to_matrix: dict[int, int] = {}

    for r in range(sheet.nrows):
        ri = sheet.rowinfo_map.get(r)
        if ri is not None and (ri.hidden or ri.height == 0):
            continue
        orig_to_matrix[r] = len(matrix)
        row: list = []
        for c in range(ncols):
            cell = sheet.cell(r, c)
            if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                row.append("")
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(int(v) if v == int(v) else v)
            else:
                row.append(str(cell.value).strip() if cell.value else "")
        matrix.append(row)
        row_heights.append(ri.height if ri is not None else 0)
        source_rows.append(r)
        if any(v != "" for v in row):
            last_filled = len(matrix) - 1

    # Обрезаем хвостовые пустые строки
    nrows_trimmed = (last_filled + 1) if last_filled >= 0 else (1 if matrix else 0)
    matrix = matrix[:nrows_trimmed]
    row_heights = row_heights[:nrows_trimmed]
    source_rows = source_rows[:nrows_trimmed]

    # Ширины столбцов
    col_widths: list[int] = []
    for c in range(ncols):
        ci = sheet.colinfo_map.get(c)
        col_widths.append(ci.width if ci is not None else 0)

    # Объединённые ячейки — перевести в пространство матрицы
    # xlrd: merged_cells = list of (row_lo, row_hi_excl, col_lo, col_hi_excl)
    merges: list[tuple[int, int, int, int]] = []
    for (row_lo, row_hi, col_lo, col_hi) in (sheet.merged_cells or []):
        visible = [orig_to_matrix[r] for r in range(row_lo, row_hi) if r in orig_to_matrix]
        if not visible:
            continue
        m_r1, m_r2 = visible[0], visible[-1]
        # Отсекаем всё, что выходит за пределы обрезанной матрицы
        if m_r1 >= nrows_trimmed:
            continue
        m_r2 = min(m_r2, nrows_trimmed - 1)
        # Переводим col_hi из «exclusive» в «inclusive»
        merges.append((m_r1, m_r2, col_lo, col_hi - 1))

    return len(matrix), ncols, matrix, last_filled, row_heights, col_widths, merges, source_rows


def load_label_template_matrix(template_path: str) -> tuple[int, int, list, int]:
    """
    Публичная обёртка для загрузки шаблона этикетки (предпросмотр в UI).
    Возвращает (nrows, ncols, matrix, last_filled_row).
    """
    nrows, ncols, matrix, last_filled, _rh, _cw, _mg, _sr = _load_template_matrix(template_path)
    return nrows, ncols, matrix, last_filled


def _write_label_block(
    ws: Any,
    matrix: list,
    template_rows: int,
    ncols: int,
    start_row: int,
    route_num: str,
    house: str,
    qty_val: Any,
    styles: dict,
    label_layout: list[dict] | None = None,
    row_heights: list[int] | None = None,
    merges: list[tuple[int, int, int, int]] | None = None,
) -> None:
    """
    Пишет один блок этикетки: копия первых template_rows строк шаблона (matrix),
    затем строка с данными (route_num / house / qty).

    row_heights — высоты строк шаблона (1/20 пункта).
    merges      — объединённые ячейки в пространстве матрицы: (r1, r2, c1, c2) включительно.
                  Ячейки внутри объединения (не верхний-левый угол) пропускаются при одиночной
                  записи; само объединение пишется через ws.write_merge().
    """
    style_yellow = styles.get("cell_yellow", styles["cell"])
    style_num = styles.get("num", styles["cell"])

    # --- Строим множество ячеек, покрытых объединением (кроме верхнего-левого угла) ---
    merged_covered: set[tuple[int, int]] = set()
    if merges:
        for (r1, r2, c1, c2) in merges:
            for mr in range(r1, r2 + 1):
                for mc in range(c1, c2 + 1):
                    if mr != r1 or mc != c1:
                        merged_covered.add((mr, mc))

    # --- Копируем строки шаблона (0 .. template_rows-1) ---
    for r in range(template_rows):
        # Применяем высоту строки из шаблона
        if row_heights and r < len(row_heights) and row_heights[r] > 0:
            ws_row = ws.row(start_row + r)
            ws_row.height = row_heights[r]
            ws_row.height_mismatch = True
        for c in range(ncols):
            if (r, c) in merged_covered:
                continue  # будет записано write_merge ниже
            val = matrix[r][c] if r < len(matrix) and c < len(matrix[r]) else ""
            cell_style = style_num if isinstance(val, (int, float)) else styles["cell"]
            try:
                if isinstance(val, (int, float)):
                    ws.write(start_row + r, c, val, cell_style)
                else:
                    ws.write(start_row + r, c, str(val), cell_style)
            except Exception as _e:
                log.debug("_write_label_block: row=%d col=%d err=%s", start_row + r, c, _e)
                ws.write(start_row + r, c, str(val), cell_style)

    # --- Записываем объединённые ячейки шаблона ---
    if merges:
        for (r1, r2, c1, c2) in merges:
            if r1 >= template_rows:
                continue
            r2_clamped = min(r2, template_rows - 1)
            val = matrix[r1][c1] if r1 < len(matrix) and c1 < len(matrix[r1]) else ""
            cell_style = style_num if isinstance(val, (int, float)) else styles["cell"]
            r1_abs = start_row + r1
            r2_abs = start_row + r2_clamped
            try:
                if r1_abs == r2_abs and c1 == c2:
                    # Одиночная ячейка — обычная запись
                    ws.write(r1_abs, c1, val if isinstance(val, (int, float)) else str(val), cell_style)
                elif isinstance(val, (int, float)):
                    ws.write_merge(r1_abs, r2_abs, c1, c2, val, cell_style)
                else:
                    ws.write_merge(r1_abs, r2_abs, c1, c2, str(val), cell_style)
            except Exception as _e:
                log.debug("_write_label_block merge r=%d-%d c=%d-%d err=%s", r1, r2, c1, c2, _e)

    # --- Заполнение данных: по layout или по умолчанию (строка данных, колонки 0,1,2) ---
    data_row = start_row + template_rows
    ncols_write = max(ncols, 3)
    if label_layout:
        values_by_cell: dict[tuple[int, int], Any] = {}
        for pl in label_layout:
            r, c = pl.get("row", template_rows), pl.get("col", 0)
            f = pl.get("field")
            if f == "routeNumber":
                val, cell_style = route_num, style_yellow
            elif f == "house":
                val, cell_style = house, styles["cell"]
            elif f == "quantity":
                val, cell_style = (qty_val if qty_val is not None else ""), style_num
            else:
                continue

            key = (r, c)
            if key in values_by_cell:
                prev_val, prev_style = values_by_cell[key]
                if prev_val and val:
                    new_val = f"{prev_val} {val}"
                else:
                    new_val = val or prev_val
                new_style = style_yellow if (prev_style is style_yellow or cell_style is style_yellow) else prev_style
                values_by_cell[key] = (new_val, new_style)
            else:
                values_by_cell[key] = (val, cell_style)

        for (r, c), (val, cell_style) in values_by_cell.items():
            try:
                row_abs = start_row + r
                if isinstance(val, (int, float)):
                    ws.write(row_abs, c, val, cell_style)
                else:
                    ws.write(row_abs, c, str(val), cell_style)
            except Exception as _e:
                log.debug("_write_label_block layout row=%d col=%d err=%s", start_row + r, c, _e)
    else:
        for c in range(ncols_write):
            if c == 0:
                val, cell_style = route_num, style_yellow
            elif c == 1:
                val, cell_style = house, styles["cell"]
            elif c == 2:
                val, cell_style = (qty_val if qty_val is not None else ""), style_num
            else:
                val, cell_style = "", styles["cell"]
            try:
                if isinstance(val, (int, float)):
                    ws.write(data_row, c, val, cell_style)
                else:
                    ws.write(data_row, c, str(val), cell_style)
            except Exception as _e:
                log.debug("_write_label_block data row col=%d err=%s", c, _e)
                ws.write(data_row, c, str(val), cell_style)


def _apply_col_widths_openpyxl(ws: Any, col_widths: list[int], ncols_cnt: int) -> None:
    """Устанавливает ширины столбцов в openpyxl Worksheet."""
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(col_widths[:ncols_cnt]):
        if w > 0:
            try:
                ws.column_dimensions[get_column_letter(ci + 1)].width = min(w / 256, 255)
            except Exception:
                pass


def _write_label_block_openpyxl(
    ws: Any,
    matrix: list,
    template_rows: int,
    ncols: int,
    start_row: int,
    route_num: str,
    house: str,
    qty_val: Any,
    label_layout: list[dict] | None = None,
    row_heights: list[int] | None = None,
    merges: list[tuple[int, int, int, int]] | None = None,
) -> None:
    """Пишет блок этикетки в openpyxl Worksheet (1-based rows/cols)."""
    from openpyxl.utils import get_column_letter
    merged_covered: set[tuple[int, int]] = set()
    if merges:
        for (r1, r2, c1, c2) in merges:
            for mr in range(r1, r2 + 1):
                for mc in range(c1, c2 + 1):
                    if mr != r1 or mc != c1:
                        merged_covered.add((mr, mc))
    for r in range(template_rows):
        if row_heights and r < len(row_heights) and row_heights[r] > 0:
            ws.row_dimensions[start_row + r + 1].height = row_heights[r] / 20.0
        for c in range(ncols):
            if (r, c) in merged_covered:
                continue
            val = matrix[r][c] if r < len(matrix) and c < len(matrix[r]) else ""
            cell = ws.cell(row=start_row + r + 1, column=c + 1, value=val if isinstance(val, (int, float)) else str(val))
    if merges:
        for (r1, r2, c1, c2) in merges:
            if r1 >= template_rows:
                continue
            r2_clamped = min(r2, template_rows - 1)
            val = matrix[r1][c1] if r1 < len(matrix) and c1 < len(matrix[r1]) else ""
            r1_abs, r2_abs = start_row + r1 + 1, start_row + r2_clamped + 1
            c1_abs, c2_abs = c1 + 1, c2 + 1
            ws.cell(row=r1_abs, column=c1_abs, value=val if isinstance(val, (int, float)) else str(val))
            if r1_abs != r2_abs or c1_abs != c2_abs:
                rng = f"{get_column_letter(c1_abs)}{r1_abs}:{get_column_letter(c2_abs)}{r2_abs}"
                ws.merge_cells(rng)
    data_row = start_row + template_rows + 1
    ncols_write = max(ncols, 3)
    if label_layout:
        values_by_cell: dict[tuple[int, int], Any] = {}
        for pl in label_layout:
            r, c = pl.get("row", template_rows), pl.get("col", 0)
            f = pl.get("field")
            if f == "routeNumber":
                val = route_num
            elif f == "house":
                val = house
            elif f == "quantity":
                val = qty_val if qty_val is not None else ""
            else:
                continue
            key = (r, c)
            if key in values_by_cell:
                prev_val = values_by_cell[key]
                val = f"{prev_val} {val}" if prev_val and val else (val or prev_val)
            values_by_cell[key] = val
        for (r, c), val in values_by_cell.items():
            ws.cell(row=start_row + r + 1, column=c + 1, value=val if isinstance(val, (int, float)) else str(val))
    else:
        ws.cell(row=data_row, column=1, value=route_num)
        ws.cell(row=data_row, column=2, value=house)
        ws.cell(row=data_row, column=3, value=qty_val if qty_val is not None else "")


def _build_label_cell_values(
    route_num: str,
    house: str,
    qty_val: Any,
    label_layout: list[dict] | None,
    template_rows: int,
) -> tuple[dict[tuple[int, int], Any], int]:
    """
    Возвращает значения для подстановки в этикетку и число добавочных строк.
    Ключ словаря: (row_idx, col_idx) в пространстве шаблона.
    """
    style_values: dict[tuple[int, int], Any] = {}
    placements = label_layout or []
    if placements:
        for pl in placements:
            r = int(pl.get("row", template_rows))
            c = int(pl.get("col", 0))
            field = pl.get("field")
            if field == "routeNumber":
                val = route_num
            elif field == "house":
                val = house
            elif field == "quantity":
                val = qty_val if qty_val is not None else ""
            else:
                continue
            key = (r, c)
            prev = style_values.get(key)
            if prev not in (None, "") and val not in (None, ""):
                style_values[key] = f"{prev} {val}"
            else:
                style_values[key] = val if val not in (None, "") else (prev or "")
    else:
        style_values = {
            (template_rows, 0): route_num,
            (template_rows, 1): house,
            (template_rows, 2): qty_val if qty_val is not None else "",
        }
    max_row_idx = max((r for r, _c in style_values), default=template_rows - 1)
    extra_rows = max(0, max_row_idx - template_rows + 1)
    return style_values, extra_rows


def _strip_windows_zone_identifier(path: str) -> None:
    """
    Удаляет ADS-метку Zone.Identifier у файла на Windows.
    Это снижает шанс открытия в Protected View и зависания при "Разрешить редактирование".
    """
    if os.name != "nt":
        return
    try:
        ads_path = f"{os.path.abspath(path)}:Zone.Identifier"
        if os.path.exists(os.path.abspath(path)):
            os.remove(ads_path)
    except Exception:
        # best-effort, генерацию не валим
        pass


def _try_generate_labels_exact_excel(
    template_path: str,
    item_list: list[tuple[str, str, float | None]],
    save_path: str,
    template_rows: int,
    source_rows: list[int],
    label_layout: list[dict] | None = None,
    output_xlsx: bool = False,
) -> bool:
    """
    Точное создание этикеток через установленный Excel в отдельном подпроцессе.
    Это защищает основное приложение от зависаний COM/скрытых диалогов Excel.
    """
    if not item_list:
        return True
    worker_path = os.path.join(os.path.dirname(__file__), "excel_exact_worker.py")
    if not os.path.isfile(worker_path):
        log.warning("Не найден worker точной генерации: %s", worker_path)
        return False

    temp_dir = tempfile.mkdtemp(prefix="labels_excel_job_")
    payload_path = os.path.join(temp_dir, "payload.json")
    try:
        with open(payload_path, "w", encoding="utf-8") as fh:
            json.dump({
                "template_path": template_path,
                "item_list": item_list,
                "save_path": save_path,
                "template_rows": template_rows,
                "source_rows": source_rows,
                "label_layout": label_layout or [],
                "output_xlsx": output_xlsx,
            }, fh, ensure_ascii=False)

        timeout_sec = max(120, min(1200, 45 + len(item_list) * 4))
        for attempt in (1, 2):
            try:
                proc = subprocess.run(
                    [sys.executable, worker_path, payload_path],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=timeout_sec,
                )
                if proc.returncode == 0 and os.path.isfile(save_path):
                    return True
                if proc.stdout.strip():
                    log.warning("Excel worker stdout (attempt %s): %s", attempt, proc.stdout.strip())
                if proc.stderr.strip():
                    log.warning("Excel worker stderr (attempt %s): %s", attempt, proc.stderr.strip())
            except subprocess.TimeoutExpired:
                log.warning(
                    "Excel worker timeout (attempt %s, timeout=%ss, items=%s)",
                    attempt, timeout_sec, len(item_list)
                )
            if attempt == 1:
                # Вторая попытка в отдельном чистом процессе Excel.
                continue
        return False
    except Exception as exc:
        log.warning("Не удалось запустить Excel worker: %s", exc)
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _try_export_xls_to_pdf(xls_path: str, pdf_path: str) -> bool:
    """Экспортирует XLS в PDF через установленный Excel в отдельном подпроцессе."""
    worker_path = os.path.join(os.path.dirname(__file__), "excel_pdf_worker.py")
    if not os.path.isfile(worker_path):
        return False
    try:
        proc = subprocess.run(
            [sys.executable, worker_path, xls_path, pdf_path],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )
        if proc.returncode == 0 and os.path.isfile(pdf_path):
            return True
        if proc.stdout.strip():
            log.warning("PDF worker stdout: %s", proc.stdout.strip())
        if proc.stderr.strip():
            log.warning("PDF worker stderr: %s", proc.stderr.strip())
        return False
    except Exception as exc:
        log.warning("Не удалось экспортировать XLS в PDF: %s", exc)
        return False


def _finalize_label_output(
    xls_path: str,
    output_format: str,
    created: list[str],
) -> None:
    """
    Добавляет итоговые файлы в список created в зависимости от output_format:
    - xls/xlsx: только xls/xlsx
    - pdf: пытается сделать pdf, при неудаче оставляет xls/xlsx
    - both: xls/xlsx + pdf (если удалось)
    """
    fmt = (output_format or "xls").lower()
    if fmt not in ("xls", "xlsx", "pdf", "both"):
        fmt = "xls"

    _strip_windows_zone_identifier(xls_path)
    if fmt in ("xls", "xlsx"):
        created.append(xls_path)
        return

    pdf_path = os.path.splitext(xls_path)[0] + ".pdf"
    if _try_export_xls_to_pdf(xls_path, pdf_path):
        _strip_windows_zone_identifier(pdf_path)
        if fmt == "pdf":
            try:
                os.remove(xls_path)
            except Exception:
                pass
            created.append(pdf_path)
            return
        created.append(xls_path)
        created.append(pdf_path)
        return

    # Если PDF не получилось — не теряем результат.
    created.append(xls_path)


def _append_labels_diagnostics(
    output_dir: str,
    lines: list[str],
) -> None:
    """Пишет диагностический отчёт генерации этикеток."""
    try:
        os.makedirs(output_dir, exist_ok=True)
        report_path = os.path.join(output_dir, "_labels_generation_report.txt")
        with open(report_path, "a", encoding="utf-8") as fh:
            fh.write("\n")
            fh.write("=" * 90 + "\n")
            fh.write(f"Запуск: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            for line in lines:
                fh.write(line.rstrip() + "\n")
    except Exception:
        pass


def _run_excel_exact_worker(payload: dict, timeout_sec: int = 180) -> tuple[int, str, str]:
    """Запускает excel_exact_worker.py с payload и возвращает (code, stdout, stderr)."""
    worker_path = os.path.join(os.path.dirname(__file__), "excel_exact_worker.py")
    if not os.path.isfile(worker_path):
        return (2, "", f"worker not found: {worker_path}")
    temp_dir = tempfile.mkdtemp(prefix="excel_exact_job_")
    payload_path = os.path.join(temp_dir, "payload.json")
    try:
        with open(payload_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False)
        proc = subprocess.run(
            [sys.executable, worker_path, payload_path],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout_sec,
        )
        return (proc.returncode, proc.stdout or "", proc.stderr or "")
    except Exception as exc:
        return (3, "", str(exc))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def get_excel_printers() -> list[str]:
    """Возвращает список доступных принтеров для печати через Excel."""
    code, out, _err = _run_excel_exact_worker({"mode": "printers"}, timeout_sec=30)
    if code != 0:
        return []
    try:
        payload = json.loads(out.strip() or "{}")
        printers = payload.get("printers") or []
        if isinstance(printers, list):
            return [str(p) for p in printers if str(p).strip()]
    except Exception:
        pass
    return []


def open_label_live_preview(xls_path: str) -> None:
    """Открывает live-preview в отдельном экземпляре Excel и ждёт закрытия книги."""
    code, _out, err = _run_excel_exact_worker(
        {"mode": "preview", "xls_path": os.path.abspath(xls_path)},
        timeout_sec=24 * 3600,
    )
    if code != 0:
        raise RuntimeError(err or "Не удалось открыть live-preview Excel.")


def print_label_file(
    xls_path: str,
    printer_name: str | None = None,
    margins: dict[str, float] | None = None,
) -> str:
    """
    Печатает файл этикеток через Excel COM.
    Возвращает фактически использованный принтер (если удалось определить).
    """
    payload = {
        "mode": "print",
        "xls_path": os.path.abspath(xls_path),
        "printer_name": printer_name or "",
        "margins": margins or {
            "top_cm": 2.0,
            "right_cm": 2.0,
            "bottom_cm": 0.0,
            "left_cm": 0.0,
        },
    }
    code, out, err = _run_excel_exact_worker(payload, timeout_sec=300)
    if code != 0:
        raise RuntimeError(err or "Не удалось отправить этикетки на печать.")
    try:
        data = json.loads((out or "").strip() or "{}")
        return str(data.get("used_printer") or printer_name or "")
    except Exception:
        return str(printer_name or "")


def export_label_to_pdf(
    xls_path: str,
    output_pdf_path: str,
    margins: dict[str, float] | None = None,
) -> str:
    """
    Экспортирует файл этикеток в PDF для точного предпросмотра.
    Возвращает путь к созданному PDF или выбрасывает RuntimeError.
    """
    margins = margins or {
        "top_cm": 2.0,
        "right_cm": 2.0,
        "bottom_cm": 0.0,
        "left_cm": 0.0,
    }
    payload = {
        "mode": "export_pdf",
        "xls_path": os.path.abspath(xls_path),
        "output_pdf_path": os.path.abspath(output_pdf_path),
        "margins": margins,
    }
    code, out, err = _run_excel_exact_worker(payload, timeout_sec=120)
    if code != 0:
        try:
            data = json.loads((err or "{}").strip() or "{}")
            msg = data.get("error", err) or "Не удалось экспортировать этикетки в PDF."
        except Exception:
            msg = err or "Не удалось экспортировать этикетки в PDF."
        raise RuntimeError(msg)
    try:
        data = json.loads((out or "").strip() or "{}")
        return str(data.get("pdf_path") or output_pdf_path)
    except Exception:
        return output_pdf_path


def prepare_label_temp_file(
    routes: list[dict],
    file_type: str,
    products_ref: list | None,
    departments_ref: list | None,
    product_name: str,
    dept_key: str | None = None,
) -> tuple[str, str]:
    """
    Готовит временный XLSX для preview/print по одному продукту.
    Формат: 3 столбца — № маршрута, Дом, Количество. Без шаблонов.
    Возвращает (xlsx_path, temp_dir). temp_dir удаляется вызывающей стороной.
    """
    from openpyxl import Workbook

    temp_dir = tempfile.mkdtemp(prefix="labels_preview_")
    save_path = os.path.join(temp_dir, f"{_safe_filename(product_name) or 'labels'}.xlsx")

    active = [r for r in routes if not r.get("excluded")]
    prod_map = {p["name"]: p for p in (products_ref or [])}
    _apply_pcs(active, prod_map)

    active_sorted = sorted(
        active,
        key=lambda r: _route_sort_key_labels(str(r.get("routeNum", ""))),
    )

    items: list[tuple[str, str, object]] = []
    for route in active_sorted:
        route_num = str(route.get("routeNum", ""))
        address = (route.get("address") or "").strip()
        house = extract_house_number(address)
        for prod in route.get("products", []):
            if prod.get("name") != product_name:
                continue
            prod_obj = prod_map.get(product_name)
            if dept_key and prod_obj and prod_obj.get("deptKey") != dept_key:
                continue
            qty = prod.get("displayQuantity", prod.get("quantity"))
            items.append((route_num, house, qty))
            break

    wb = Workbook()
    ws = wb.active
    ws.title = "Этикетки"
    ws.cell(row=1, column=1, value="№ маршрута")
    ws.cell(row=1, column=2, value="Дом")
    ws.cell(row=1, column=3, value="Количество")
    for row_idx, (route_num, house, qty) in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=route_num)
        ws.cell(row=row_idx, column=2, value=house)
        ws.cell(row=row_idx, column=3, value=qty if qty is not None else "")
    wb.save(save_path)

    return (save_path, temp_dir)


def _include_product_for_labels(
    prod_name: str,
    products_ref: list | None,
    departments_ref: list | None,
    file_type: str,
    only_dept_key: str | None = None,
    exclude_chistchenka_dirty: bool = False,
) -> bool:
    """
    Общая логика: включать ли продукт в этикетки.
    Возвращает False для продуктов с неизвестным отделом (сиротские продукты).
    exclude_chistchenka_dirty: исключать продукты «В Грязные» подотдела Чищенка.
    """
    if not products_ref or not departments_ref:
        return True
    prod = next((p for p in products_ref if p.get("name") == prod_name), None)
    if not prod:
        return True
    if exclude_chistchenka_dirty and prod.get("showInDirty") and data_store.is_subdept_chistchenka(prod.get("deptKey")):
        return False
    dept_key = prod.get("deptKey")
    if not dept_key:
        return False
    if only_dept_key and dept_key != only_dept_key:
        return False
    for dept in departments_ref:
        if dept.get("key") == dept_key:
            if not dept.get("labelsEnabled", True):
                return False
            if dept.get("labelsFor", "both") in ("both", file_type):
                return True
            return False
        for sub in dept.get("subdepts", []):
            if sub.get("key") == dept_key:
                if not sub.get("labelsEnabled", True):
                    return False
                if sub.get("labelsFor", "both") in ("both", file_type):
                    return True
                return False
    return False  # dept_key не найден в departments — исключаем


def _products_for_labels_simple(
    products_ref: list | None,
    departments_ref: list | None,
    file_type: str,
    only_product: str | None = None,
    only_dept_key: str | None = None,
) -> dict[str, str]:
    """Продукты для этикеток (без шаблонов): {prod_name: dept_key}."""
    result: dict[str, str] = {}
    for p in products_ref or []:
        name = p.get("name", "")
        if not name:
            continue
        if only_product and name != only_product:
            continue
        if _include_product_for_labels(name, products_ref, departments_ref, file_type, only_dept_key):
            result[name] = p.get("deptKey") or ""
    return result


def labels_preview(
    routes: list[dict],
    file_type: str,
    products_ref: list | None,
    departments_ref: list | None,
    only_product: str | None = None,
    only_dept_key: str | None = None,
) -> list[tuple[str, str, int]]:
    """
    Возвращает список (продукт, отдел, кол-во маршрутов) для предпросмотра этикеток.
    Без шаблонов — все продукты отделов с labelsEnabled.
    """
    active = [r for r in routes if not r.get("excluded")]
    products_map = _products_for_labels_simple(
        products_ref, departments_ref, file_type, only_product, only_dept_key
    )

    result: list[tuple[str, str, int]] = []
    for prod_name, dept_key in products_map.items():
        route_nums: set[str] = set()
        for route in active:
            for prod in route.get("products", []):
                if prod.get("name") == prod_name:
                    route_nums.add(str(route.get("routeNum", "")))
                    break
        if not route_nums:
            continue
        dept_name = _dept_display_name(dept_key, departments_ref)
        result.append((prod_name, dept_name, len(route_nums)))
    return sorted(result, key=lambda x: (x[1], x[0]))


def labels_preview_rows(
    routes: list[dict],
    file_type: str,
    products_ref: list | None,
    departments_ref: list | None,
    only_product: str | None = None,
    only_dept_key: str | None = None,
) -> list[tuple[str, str, str, str, str]]:
    """
    Возвращает список строк этикеток для таблицы предпросмотра.
    Каждая строка: (№ маршрута, Адрес, Продукт, Отдел, Кол-во).
    """
    active = [r for r in routes if not r.get("excluded")]
    products_map = _products_for_labels_simple(
        products_ref, departments_ref, file_type, only_product, only_dept_key
    )

    prod_map = {p["name"]: p for p in (products_ref or [])}
    _apply_pcs(active, prod_map)

    active_sorted = sorted(
        active,
        key=lambda r: _route_sort_key_labels(str(r.get("routeNum", ""))),
    )

    rows: list[tuple[str, str, str, str, str]] = []
    for route in active_sorted:
        route_num = str(route.get("routeNum", ""))
        address = (route.get("address") or "").strip()
        for prod in route.get("products", []):
            prod_name = prod.get("name", "")
            if prod_name not in products_map:
                continue
            dept_key = products_map[prod_name]
            dept_name = _dept_display_name(dept_key, departments_ref)
            qty = prod.get("displayQuantity", prod.get("quantity"))
            qty_str = str(qty) if qty is not None else ""
            rows.append((route_num, address, prod_name, dept_name, qty_str))

    return rows


def generate_simple_labels(
    routes: list[dict],
    base_dir: str,
    file_type: str,
    products_ref: list | None,
    departments_ref: list | None,
    date_str: str | None = None,
) -> list[str]:
    """
    Создаёт файлы этикеток без шаблонов.
    Формат: 3 столбца — № маршрута, Дом (из адреса), Количество.
    Файлы: base_dir/.../Основные|Увеличение/этикетки/{отдел}/{продукт}.xlsx
    """
    from openpyxl import Workbook

    active = [r for r in routes if not r.get("excluded")]
    if not active:
        return []

    date_part = date_str or get_routes_date_str()
    type_dir = get_routes_type_folder(base_dir, file_type, date_str)
    labels_dir = os.path.join(type_dir, f"Этикетки на {date_part}")
    os.makedirs(labels_dir, exist_ok=True)

    products_map = _products_for_labels_simple(products_ref, departments_ref, file_type, None, None)
    if not products_map:
        return []

    created: list[str] = []
    prod_map = {p["name"]: p for p in (products_ref or [])}
    _apply_pcs(active, prod_map)

    active_sorted = sorted(
        active,
        key=lambda r: _route_sort_key_labels(str(r.get("routeNum", ""))),
    )

    by_product_dept: dict[tuple[str, str], list[tuple[str, str, object]]] = {}
    for route in active_sorted:
        route_num = str(route.get("routeNum", ""))
        address = (route.get("address") or "").strip()
        house = extract_house_number(address)
        for prod in route.get("products", []):
            prod_name = prod.get("name", "")
            if prod_name not in products_map:
                continue
            dept_key = products_map[prod_name]
            dept_display = _dept_display_name(dept_key, departments_ref)
            if not dept_display:
                dept_display = "Прочее"
            qty = prod.get("displayQuantity", prod.get("quantity"))
            key = (prod_name, dept_display)
            if key not in by_product_dept:
                by_product_dept[key] = []
            by_product_dept[key].append((route_num, house, qty))

    for (prod_name, dept_display), items in by_product_dept.items():
        if not items:
            continue
        safe_dept = _safe_filename(dept_display) or "Отдел"
        safe_prod = _safe_filename(prod_name) or "Продукт"
        dept_folder = os.path.join(labels_dir, safe_dept)
        os.makedirs(dept_folder, exist_ok=True)
        save_path = os.path.join(dept_folder, f"{safe_prod} {date_part}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Этикетки"
        ws.cell(row=1, column=1, value="№ маршрута")
        ws.cell(row=1, column=2, value="Дом")
        ws.cell(row=1, column=3, value="Количество")
        for row_idx, (route_num, house, qty) in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=route_num)
            ws.cell(row=row_idx, column=2, value=house)
            ws.cell(row=row_idx, column=3, value=qty if qty is not None else "")
        wb.save(save_path)
        created.append(save_path)

    return created


def generate_labels_from_templates(
    routes: list[dict],
    output_dir: str,
    file_type: str,
    products_ref: list | None,
    departments_ref: list | None,
    only_product: str | None = None,
    only_dept_key: str | None = None,
    dept_subfolders: bool = False,
    overwrite: bool = True,
    output_format: str = "xlsx",
    strict_exact: bool = True,
    diagnostics_dir: str | None = None,
) -> list[str]:
    """
    Создаёт этикетки XLS по шаблонам продуктов.
    Один файл на продукт (или два для сыпучки: до 4 кг / после 4 кг). Количество сравнивается как float.
    Учитывается labelsEnabled отдела/подотдела. Режимы: default; chistchenka (макс. кг на этикетку, дубликаты пока не останется < макс — последняя этикетка; файл: продукт_дата_основной/увеличение);
    сыпучка — два файла до/после порога.
    only_product / only_dept_key — генерация только для одного продукта или отдела.
    """
    active = [r for r in routes if not r.get("excluded")]
    created: list[str] = []
    styles = _get_styles()
    diagnostics: list[str] = []

    active_sorted = sorted(
        active,
        key=lambda r: _route_sort_key_labels(str(r.get("routeNum", "")))
    )

    prod_map = {p["name"]: p for p in (products_ref or [])}
    _apply_pcs(active_sorted, prod_map)

    # Продукты с недоступным шаблоном — в диагностику
    for p in products_ref or []:
        tpl = p.get("labelTemplatePath") or ""
        if tpl and not os.path.isfile(tpl):
            diagnostics.append(f"[WARN] {p.get('name', '?')}: шаблон недоступен — {tpl}")

    products_in_routes: set[str] = set()
    for r in active:
        for prod in r.get("products", []):
            name = prod.get("name")
            if name:
                products_in_routes.add(name)

    products_with_templates: dict[str, tuple[str, str, list]] = {}
    for p in products_ref or []:
        tpl = p.get("labelTemplatePath") or ""
        if not tpl or not os.path.isfile(tpl):
            continue
        name = p.get("name", "")
        if name not in products_in_routes:
            continue
        if only_product and name != only_product:
            continue
        if _include_product_for_labels(name, products_ref, departments_ref, file_type, only_dept_key, exclude_chistchenka_dirty=True):
            layout = p.get("labelLayout")  # list of {row, col, field} or None
            if layout and not isinstance(layout, list):
                layout = None
            products_with_templates[name] = (tpl, p.get("deptKey") or "", layout or [])

    def _apply_col_widths(ws_sheet: Any, col_widths: list[int], ncols_cnt: int) -> None:
        """Устанавливает ширины столбцов из шаблона на листе."""
        for ci, w in enumerate(col_widths[:ncols_cnt]):
            if w > 0:
                ws_sheet.col(ci).width = w

    def _write_product_labels(
        item_list: list[tuple[str, str, float | None]],
        save_path: str,
        template_rows: int,
        ncols: int,
        matrix: list,
        label_layout: list[dict] | None = None,
        row_heights: list[int] | None = None,
        col_widths: list[int] | None = None,
        merges: list[tuple[int, int, int, int]] | None = None,
    ) -> None:
        wb = xlwt.Workbook(encoding="utf-8")
        block_height = template_rows + 1  # строки шаблона + одна строка с данными
        _XLS_MAX_ROWS = 65535

        sheet_num = 1
        ws = wb.add_sheet("Этикетки")
        if col_widths:
            _apply_col_widths(ws, col_widths, ncols)
        page_breaks: list[int] = []
        row = 0

        for route_num, house, qty in item_list:
            # Если следующий блок не помещается на текущий лист — создаём новый
            if row + block_height > _XLS_MAX_ROWS:
                if page_breaks:
                    ws.horz_page_breaks = [(r, 0, 255) for r in page_breaks[:-1]]
                sheet_num += 1
                ws = wb.add_sheet(f"Этикетки {sheet_num}")
                if col_widths:
                    _apply_col_widths(ws, col_widths, ncols)
                page_breaks = []
                row = 0

            _write_label_block(
                ws, matrix, template_rows, ncols, row,
                route_num, house, qty, styles, label_layout, row_heights, merges,
            )
            row += block_height
            page_breaks.append(row)

        if page_breaks:
            ws.horz_page_breaks = [(r, 0, 255) for r in page_breaks[:-1]]
        wb.save(save_path)

    def _write_product_labels_xlsx(
        item_list: list[tuple[str, str, float | None]],
        save_path: str,
        template_rows: int,
        ncols: int,
        matrix: list,
        label_layout: list[dict] | None = None,
        row_heights: list[int] | None = None,
        col_widths: list[int] | None = None,
        merges: list[tuple[int, int, int, int]] | None = None,
    ) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        block_height = template_rows + 1
        _XLSX_MAX_ROWS = 1048576

        sheet_num = 1
        ws = wb.active
        ws.title = "Этикетки"
        if col_widths:
            _apply_col_widths_openpyxl(ws, col_widths, ncols)
        row = 0

        for route_num, house, qty in item_list:
            if row + block_height > _XLSX_MAX_ROWS:
                sheet_num += 1
                ws = wb.create_sheet(f"Этикетки {sheet_num}")
                if col_widths:
                    _apply_col_widths_openpyxl(ws, col_widths, ncols)
                row = 0

            _write_label_block_openpyxl(
                ws, matrix, template_rows, ncols, row,
                route_num, house, qty, label_layout, row_heights, merges,
            )
            row += block_height

        wb.save(save_path)

    try:
        for prod_name, (template_path, dept_key, label_layout) in products_with_templates.items():
            started = time.perf_counter()
            mode = _label_print_mode_for_dept(dept_key, departments_ref)
            label_rules = _label_rules_for_dept(dept_key, departments_ref)
            try:
                nrows, ncols, matrix, last_filled, row_heights, col_widths, merges, source_rows = _load_template_matrix(template_path)
            except Exception as _e:
                log.warning("Не удалось загрузить шаблон '%s': %s", template_path, _e)
                diagnostics.append(f"[ERROR] {prod_name}: шаблон не прочитан ({_e})")
                continue

            template_rows = nrows if nrows > 0 else 1
            ncols = max(ncols, 3)

            items: list[tuple[str, str, float | None]] = []
            for route in active_sorted:
                route_num = str(route.get("routeNum", ""))
                house = extract_house_number(route.get("address", ""))
                for prod in route.get("products", []):
                    if prod.get("name") != prod_name:
                        continue
                    items.append((route_num, house, prod.get("displayQuantity", prod.get("quantity"))))

            if not items:
                diagnostics.append(f"[SKIP] {prod_name}: нет данных для генерации")
                continue

            safe_name = _safe_filename(prod_name)

            if dept_subfolders and dept_key:
                dept_display = _dept_display_name(dept_key, departments_ref)
                safe_dept = _safe_filename(dept_display) if dept_display else "Отдел"
                actual_dir = os.path.join(output_dir, safe_dept)
            else:
                actual_dir = output_dir
            os.makedirs(actual_dir, exist_ok=True)

            use_xlsx = (output_format or "xls").lower() == "xlsx"
            ext = ".xlsx" if use_xlsx else ".xls"

            def _unique_path(base_path: str, base_name: str, ext_suffix: str = ext) -> str:
                if overwrite:
                    return base_path
                cnt = 0
                p = base_path
                while os.path.exists(p):
                    cnt += 1
                    p = os.path.join(actual_dir, f"{base_name}_{cnt}{ext_suffix}")
                return p

            try:
                if mode == "sypuchka":
                    syp = label_rules.get("sypuchka") or {}
                    threshold = float(syp.get("thresholdKg", 4))
                    label_below = syp.get("labelBelow", "меньше 4 кг")
                    label_above = syp.get("labelAbove", "больше 4 кг")
                    items_below = [(rn, h, q) for rn, h, q in items if q is not None and float(q) <= threshold]
                    items_above = [(rn, h, q) for rn, h, q in items if q is None or float(q) > threshold]
                    for title_suffix, item_list in [(label_below, items_below), (label_above, items_above)]:
                        if not item_list:
                            continue
                        base_name = f"Этикетки {title_suffix} {safe_name}"
                        save_path = _unique_path(os.path.join(actual_dir, f"{base_name}{ext}"), base_name)
                        exact_ok = _try_generate_labels_exact_excel(
                            template_path, item_list, save_path, template_rows, source_rows, label_layout,
                            output_xlsx=use_xlsx,
                        )
                        if not exact_ok and strict_exact:
                            raise RuntimeError(
                                f"Excel COM не смог создать этикетки для '{prod_name}' ({title_suffix})."
                            )
                        if not exact_ok:
                            if use_xlsx:
                                _write_product_labels_xlsx(
                                    item_list, save_path, template_rows, ncols, matrix,
                                    label_layout, row_heights, col_widths, merges,
                                )
                            else:
                                _write_product_labels(
                                    item_list, save_path, template_rows, ncols, matrix,
                                    label_layout, row_heights, col_widths, merges,
                                )
                        _finalize_label_output(save_path, output_format, created)

                elif mode == "chistchenka":
                    ch = label_rules.get("chistchenka") or {}
                    max_kg = float(ch.get("maxKgPerLabel", 5))
                    if max_kg <= 0:
                        max_kg = 5.0
                    expanded: list[tuple[str, str, float]] = []
                    for route_num, house, qty in items:
                        try:
                            val = float(qty) if qty is not None else 0.0
                        except (TypeError, ValueError):
                            val = 0.0
                        while val > 0:
                            take = min(max_kg, val)
                            expanded.append((route_num, house, take))
                            val -= take
                    date_str = _format_date(date.today())
                    type_lbl_short = "основной" if file_type == "main" else "увеличение"
                    base_name = f"{safe_name}_{date_str}_{type_lbl_short}"
                    save_path = _unique_path(os.path.join(actual_dir, f"{base_name}{ext}"), base_name)
                    exact_ok = _try_generate_labels_exact_excel(
                        template_path, expanded, save_path, template_rows, source_rows, label_layout,
                        output_xlsx=use_xlsx,
                    )
                    if not exact_ok and strict_exact:
                        raise RuntimeError(
                            f"Excel COM не смог создать этикетки для '{prod_name}'."
                        )
                    if not exact_ok:
                        if use_xlsx:
                            _write_product_labels_xlsx(
                                expanded, save_path, template_rows, ncols, matrix,
                                label_layout, row_heights, col_widths, merges,
                            )
                        else:
                            _write_product_labels(
                                expanded, save_path, template_rows, ncols, matrix,
                                label_layout, row_heights, col_widths, merges,
                            )
                    _finalize_label_output(save_path, output_format, created)

                else:
                    date_str = _format_date(date.today())
                    type_lbl_short = "основной" if file_type == "main" else "увеличение"
                    base_name = f"{safe_name}_{date_str}_{type_lbl_short}"
                    save_path = _unique_path(os.path.join(actual_dir, f"{base_name}{ext}"), base_name)
                    exact_ok = _try_generate_labels_exact_excel(
                        template_path, items, save_path, template_rows, source_rows, label_layout,
                        output_xlsx=use_xlsx,
                    )
                    if not exact_ok and strict_exact:
                        raise RuntimeError(
                            f"Excel COM не смог создать этикетки для '{prod_name}'."
                        )
                    if not exact_ok:
                        if use_xlsx:
                            _write_product_labels_xlsx(
                                items, save_path, template_rows, ncols, matrix,
                                label_layout, row_heights, col_widths, merges,
                            )
                        else:
                            _write_product_labels(
                                items, save_path, template_rows, ncols, matrix,
                                label_layout, row_heights, col_widths, merges,
                            )
                    _finalize_label_output(save_path, output_format, created)

                elapsed = int((time.perf_counter() - started) * 1000)
                diagnostics.append(
                    f"[OK] {prod_name}: mode={mode}, items={len(items)}, time={elapsed}ms"
                )
            except Exception as exc:
                elapsed = int((time.perf_counter() - started) * 1000)
                diagnostics.append(
                    f"[ERROR] {prod_name}: mode={mode}, time={elapsed}ms, error={exc}"
                )
                log.warning("Ошибка генерации этикеток для '%s': %s", prod_name, exc)
                continue
    finally:
        _append_labels_diagnostics(output_dir, diagnostics)
        if diagnostics_dir and os.path.abspath(diagnostics_dir) != os.path.abspath(output_dir):
            _append_labels_diagnostics(diagnostics_dir, diagnostics)

    return created


def _get_pcs_unit_label(dept_key: str | None) -> str:
    """Единица для столбца шт: «кор.» для полуфабрикатов, иначе «шт»."""
    if dept_key and data_store.is_subdept_polufabricates(dept_key):
        return "кор."
    return "шт"


def _fmt_pcs_int(val) -> str | None:
    """Возвращает шт как целое число для отображения, иначе None."""
    if val is None:
        return None
    try:
        v = float(val)
        return str(int(round(v)))
    except (TypeError, ValueError):
        return str(val) if val != "" else None


def _fmt_pcs_cell(prod: dict, dept_key: str | None = None) -> str:
    """Форматирует шт для отдельной ячейки: целое число или «pcs кор./шт + tail unit» для полуфабрикатов.
    При 0 шт + хвостик возвращает пустую строку (количество уже в столбце quantity)."""
    pcs = prod.get("pcs")
    if pcs is None:
        return ""
    pi = _fmt_pcs_int(pcs)
    tail = prod.get("pcsTail")
    unit = (prod.get("unit") or "").strip()
    pcs_unit = _get_pcs_unit_label(dept_key)
    if tail is not None and tail > 1e-9 and unit:
        if int(pcs or 0) == 0:
            return ""
        try:
            tval = float(tail)
            tail_txt = str(int(tval)) if abs(tval - round(tval)) < 1e-9 else f"{tval:.3f}".rstrip("0").rstrip(".")
            return f"{pi} {pcs_unit} + {tail_txt} {unit}"
        except (TypeError, ValueError):
            pass
    return f"{pi} {pcs_unit}" if pi else str(pcs) if pcs is not None else ""


def _fmt_qty_with_pcs(prod: dict, dept_key: str | None = None) -> str:
    """Форматирует количество с опциональным значением шт/кор.
    Использует displayQuantity (уже с учётом множителя замены), если задано.
    Формат: «количество / шт» или «количество / pcs кор./шт + tail unit» для полуфабрикатов.
    При 0 шт + хвостик — только количество (нет смысла в «0 шт + хвостик»).
    Для объединённых замен (_merged): displayQuantity и pcs уже отформатированы.
    """
    if prod.get("_merged"):
        q = prod.get("displayQuantity") or prod.get("quantity") or ""
        pc = prod.get("pcs") or prod.get("pcs_display")
        return f"{q} / {pc}" if pc else str(q)
    qty = prod.get("displayQuantity", prod.get("quantity"))
    unit = (prod.get("unit") or "").strip()
    pcs = prod.get("pcs")
    tail = prod.get("pcsTail")
    if qty is None:
        return ""
    qty_str = f"{qty} {unit}".strip() if unit else str(qty)
    pcs_int = _fmt_pcs_int(pcs)
    pcs_unit = _get_pcs_unit_label(dept_key)
    # Полуфабрикаты: слева количество, справа «pcs кор./шт + tail unit» только если pcs > 0
    if pcs_int is not None and tail is not None and tail > 1e-9 and unit:
        try:
            tval = float(tail)
            tail_txt = str(int(tval)) if abs(tval - round(tval)) < 1e-9 else f"{tval:.3f}".rstrip("0").rstrip(".")
            pcs_tail_str = f"{pcs_int} {pcs_unit} + {tail_txt} {unit}"
            if int(pcs or 0) > 0:
                return f"{qty_str} / {pcs_tail_str}"
            return qty_str
        except (ValueError, TypeError):
            pass
    if pcs_int is not None and (tail is None or tail <= 1e-9):
        return f"{qty_str} / {pcs_int} {pcs_unit}"
    return qty_str


# ─────────────────────────── Замена продукта ───────────────────────────────

def merge_replacement_pairs_for_display(
    products: list[dict],
    replacements: list[dict],
) -> list[dict]:
    """
    Объединяет пары (fromProduct, toProduct) из замен в одну строку для отображения.
    Когда в маршруте есть и основной продукт, и замена — показываем «Молоко + Кефир (замена)» с «7 + 3» и «7 шт + 3 шт».
    """
    if not replacements:
        return list(products)
    result: list[dict] = []
    merged_names: set[str] = set()

    for repl in replacements:
        from_name = repl.get("fromProduct") or ""
        to_products = repl.get("toProducts")
        if to_products and len(to_products) >= 2:
            to_names = list(to_products)[:2]
        else:
            to_name = repl.get("toProduct") or ""
            if not to_name:
                continue
            to_names = [to_name]
        if not from_name or not to_names or from_name in to_names:
            continue
        from_p = next((p for p in products if p.get("name") == from_name), None)
        to_ps = [next((p for p in products if p.get("name") == n), None) for n in to_names]
        if not from_p or not all(to_ps) or from_name in merged_names or any(n in merged_names for n in to_names):
            continue
        unit = (from_p.get("unit") or (to_ps[0].get("unit") if to_ps else "") or "").strip()
        qtys = [from_p.get("displayQuantity", from_p.get("quantity"))]
        pcs_vals = [from_p.get("pcs")]
        for tp in to_ps:
            qtys.append(tp.get("displayQuantity", tp.get("quantity")))
            pcs_vals.append(tp.get("pcs"))
        qty_str = _fmt_merged_qty_multi(qtys, unit)
        pcs_str = _fmt_merged_pcs_multi(pcs_vals)
        names_display = " + ".join([from_name] + to_names)
        result.append({
            "name": f"{names_display} (замена)",
            "unit": unit,
            "displayQuantity": qty_str,
            "quantity": qty_str,
            "pcs": pcs_str,
            "pcs_display": pcs_str,
            "_merged": True,
        })
        merged_names.add(from_name)
        merged_names.update(to_names)

    for p in products:
        if p.get("name") not in merged_names:
            result.append(dict(p))

    return result


def _fmt_merged_qty_multi(qtys: list, unit: str) -> str:
    """Форматирует объединённое количество для нескольких значений: «7 + 3 + 2» или «7 + 3 + 2 л»."""
    def _v(x):
        if x is None:
            return ""
        try:
            v = float(x)
            return str(int(v)) if abs(v - round(v)) < 1e-9 else f"{v:.2f}".rstrip("0").rstrip(".")
        except (TypeError, ValueError):
            return str(x)
    parts = [_v(q) for q in qtys if _v(q)]
    if not parts:
        return ""
    return " + ".join(parts) + (f" {unit}" if unit else "")


def _fmt_merged_pcs_multi(pcs_vals: list) -> str | None:
    """Форматирует объединённые шт для нескольких значений: «7 шт + 3 шт + 2 шт»."""
    vv = [_fmt_pcs_int(p) for p in pcs_vals]
    vv = [v for v in vv if v]
    if not vv:
        return None
    return " + ".join(f"{v} шт" for v in vv)


def _fmt_merged_qty(q1, q2, unit: str) -> str:
    """Форматирует объединённое количество: «7 + 3» или «7 + 3 л»."""
    def _v(x):
        if x is None:
            return ""
        try:
            v = float(x)
            return str(int(v)) if abs(v - round(v)) < 1e-9 else f"{v:.2f}".rstrip("0").rstrip(".")
        except (TypeError, ValueError):
            return str(x)
    s1, s2 = _v(q1), _v(q2)
    if not s1 and not s2:
        return ""
    if not s1:
        return f"{s2} {unit}".strip() if unit else s2
    if not s2:
        return f"{s1} {unit}".strip() if unit else s1
    return f"{s1} + {s2}" + (f" {unit}" if unit else "")


def _fmt_merged_pcs(pcs1, pcs2) -> str | None:
    """Форматирует объединённые шт: «7 шт + 3 шт». Шт — всегда целое число."""
    v1, v2 = _fmt_pcs_int(pcs1), _fmt_pcs_int(pcs2)
    if v1 is None and v2 is None:
        return None
    if v1 is None:
        return f"{v2} шт" if v2 else None
    if v2 is None:
        return f"{v1} шт"
    return f"{v1} шт + {v2} шт"


def apply_replacements(
    routes: list[dict],
    replacements: list[dict],
    sort_asc: bool = True,
) -> list[dict]:
    """
    Применяет замены продуктов к маршрутам. Возвращает копию routes с заменёнными продуктами.

    full: во всех маршрутах fromProduct → toProduct (то же количество).
    partial: в выбранных учреждениях, с последних по порядку маршрутов, replace quantity.
    """
    if not replacements:
        return routes
    data = copy.deepcopy(routes)

    for repl in replacements:
        from_name = repl.get("fromProduct") or ""
        to_products = repl.get("toProducts")
        if to_products and len(to_products) >= 2:
            to_names = list(to_products)[:2]
            split_ratio = float(repl.get("splitRatio", 0.5) or 0.5)
            split_ratio = max(0, min(1, split_ratio))
        else:
            to_name = repl.get("toProduct") or ""
            if not to_name or from_name == to_name:
                continue
            to_names = [to_name]
            split_ratio = 1.0
        if not from_name or not to_names:
            continue
        if from_name in to_names:
            continue
        mode = repl.get("mode", "full")
        unit = repl.get("unit", "")

        def _add_or_merge(prods: list, target_name: str, add_qty: float) -> None:
            found = next((p for p in prods if p.get("name") == target_name), None)
            if found:
                found["quantity"] = float(found.get("quantity") or 0) + add_qty
            else:
                prods.append({"name": target_name, "quantity": add_qty, "unit": unit})

        if mode == "full":
            if len(to_names) == 1:
                for route in data:
                    for prod in route.get("products", []):
                        if prod.get("name") == from_name:
                            prod["name"] = to_names[0]
            else:
                for route in data:
                    for prod in route.get("products", []):
                        if prod.get("name") != from_name:
                            continue
                        try:
                            qty = float(prod.get("quantity") or 0)
                        except (TypeError, ValueError):
                            qty = 0
                        if qty <= 0:
                            continue
                        q1 = qty * split_ratio
                        q2 = qty * (1 - split_ratio)
                        prods = route.get("products", [])
                        prod["name"] = to_names[0]
                        prod["quantity"] = q1
                        _add_or_merge(prods, to_names[1], q2)

        else:  # partial
            qty_to_replace = float(repl.get("quantity", 0) or 0)
            if qty_to_replace <= 0:
                continue
            addresses = set(a for a in (repl.get("addresses") or []) if a)
            inst_codes = set(repl.get("institutionCodes") or [])
            if addresses:
                filtered = [r for r in data if (r.get("address") or "").strip() in addresses]
            elif inst_codes:
                filtered = [
                    r for r in data
                    if data_store.get_institution_key_from_address(r.get("address") or "") in inst_codes
                ]
            else:
                continue
            sorted_routes = _sort_routes(filtered, sort_asc)
            sorted_routes = list(reversed(sorted_routes))

            remaining = qty_to_replace
            for route in sorted_routes:
                if remaining <= 0:
                    break
                prods = route.get("products", [])
                # Собираем кандидатов (fromProduct) с конца списка продуктов
                candidates: list[tuple[int, dict, float]] = []
                for i, prod in enumerate(prods):
                    if prod.get("name") != from_name:
                        continue
                    try:
                        qty = float(prod.get("quantity") or 0)
                    except (TypeError, ValueError):
                        qty = 0
                    if qty > 0:
                        candidates.append((i, prod, qty))
                # Обрабатываем с конца (последние маршруты = последние продукты в списке)
                for i, prod, qty in reversed(candidates):
                    if remaining <= 0:
                        break
                    take = min(remaining, qty)
                    remaining -= take

                    take1 = take * split_ratio
                    take2 = take * (1 - split_ratio)
                    if abs(qty - take) < 1e-9:
                        prod["name"] = to_names[0]
                        prod["quantity"] = take1
                        if len(to_names) >= 2 and take2 > 1e-9:
                            _add_or_merge(prods, to_names[1], take2)
                    else:
                        prod["quantity"] = qty - take
                        _add_or_merge(prods, to_names[0], take1)
                        if len(to_names) >= 2 and take2 > 1e-9:
                            _add_or_merge(prods, to_names[1], take2)

    return data


# ─────────────────────────── Общие маршруты ───────────────────────────────

def generate_general_routes(
    routes: list[dict],
    file_type: str,
    save_path: str,
    products_settings: dict[str, dict],
    sort_asc: bool = True,
    date_str: str | None = None,
    replacements: list[dict] | None = None,
) -> str:
    """
    Создаёт файл «Общие маршруты».
    Один лист, все маршруты подряд с разрывом страницы перед каждым следующим блоком.

    Структура блока на маршрут:
      Строка 1: дата + тип файла (заголовок)
      Строка 2: «№ маршрута» | «Адрес»
      Строка 3: заголовки столбцов продуктов
      Строки 4+: данные продуктов

    Маршруты отсортированы по возрастанию номера (или по sort_asc).
    """
    routes = apply_replacements(routes, replacements or [], sort_asc)
    _apply_pcs(routes, products_settings)
    routes_sorted = _sort_routes(routes, sort_asc)

    # Для одного листа — общее число столбцов по всем маршрутам (если хоть у одного есть шт)
    has_pcs_any = any(
        any(p.get("pcs") is not None for p in r.get("products", []))
        for r in routes_sorted
    )
    n_data_cols = 5 if has_pcs_any else 4

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Маршруты")
    _apply_page_margins(ws, for_labels=False)
    styles = _get_styles()

    date_str = date_str or _format_date(_tomorrow())
    type_lbl = _type_label(file_type)
    header_text = f"{date_str}  {type_lbl}"

    prod_headers = ["Продукт", "Ед. изм.", "Количество"]
    if has_pcs_any:
        prod_headers.append("Шт")

    page_breaks: list[int] = []
    row = 0

    for route in routes_sorted:
        if row > 0:
            page_breaks.append(row)  # разрыв страницы перед блоком маршрута

        products = route.get("products", [])
        has_pcs = has_pcs_any  # используем общие столбцы
        n_prods = len(products)

        # Строка 1: дата + тип файла
        ws.write_merge(row, row, 0, n_data_cols - 1, header_text, styles["title"])
        row += 1

        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")
        ws.write(row, 0, route_num_str, styles["header"])
        if n_data_cols > 2:
            ws.write_merge(row, row, 1, n_data_cols - 1, address, styles["header_wrap"])
        else:
            ws.write(row, 1, address, styles["header_wrap"])
        row += 1

        ws.write(row, 0, "", styles["header"])
        for ci, h in enumerate(prod_headers):
            ws.write(row, 1 + ci, h, styles["header"])
        row += 1

        if n_prods == 0:
            ws.write(row, 0, "", styles["cell"])
            ws.write(row, 1, "", styles["cell"])
            ws.write(row, 2, "", styles["cell"])
            ws.write(row, 3, "", styles["num"])
            if has_pcs_any:
                ws.write(row, 4, "", styles["num"])
            row += 1
        else:
            display_prods = merge_replacement_pairs_for_display(
                products, replacements or []
            )
            for prod in display_prods:
                ws.write(row, 0, "", styles["cell"])
                ws.write(row, 1, prod.get("name", ""), styles["cell"])
                ws.write(row, 2, prod.get("unit", ""), styles["cell"])
                qty = prod.get("displayQuantity", prod.get("quantity"))
                ws.write(row, 3, qty if qty is not None and qty != "" else "", styles["num"])
                if has_pcs_any:
                    pcs = prod.get("pcs_display") if prod.get("_merged") else _fmt_pcs_cell(prod)
                    ws.write(row, 4, pcs if pcs else "", styles["num"])
                row += 1

        # Хвостовая строка: № + номер маршрута + "    " + дом/строение/корпус (жирный 30pt)
        house_part = _extract_house_parts(address, include_route_num=False)
        if not house_part:
            house_part = (address or "").strip()
        tail_text = f"{ROUTE_SIGN}{route_num_str}    {house_part}"
        ws.write_merge(row, row, 0, n_data_cols - 1, tail_text, styles["tail_line"])
        row += 1

    if page_breaks:
        ws.horz_page_breaks = [(r, 0, 255) for r in page_breaks]

    _set_col_width(ws, 0, 14)
    _set_col_width(ws, 1, 42)
    _set_col_width(ws, 2, 12)
    _set_col_width(ws, 3, 14)
    if has_pcs_any:
        _set_col_width(ws, 4, 8)

    _safe_save_workbook(wb, save_path)
    return save_path


# ─────────────────────────── Форматы файлов по отделам ────────────────────

def _write_dept_wide(
    ws: xlwt.Worksheet,
    routes: list[dict],
    dept_name: str,
    date_str: str,
    type_lbl: str,
    styles: dict[str, xlwt.XFStyle],
    sort_asc: bool = False,
) -> None:
    """
    Формат 1 (wide): каждый уникальный продукт — отдельный столбец.

    Структура:
      Строка 1: заголовок (объединённая)
      Строка 2: Маршрут | Адрес | ПродуктA | ПродуктB | ...
      Строка 3+: данные маршрутов (одна строка на маршрут)

    Значение в ячейке продукта: "5 кг / 3 шт" (шт — опционально).
    """
    # Собираем уникальные продукты в порядке первого появления
    unique_prods: list[str] = []
    seen: set[str] = set()
    for route in routes:
        for prod in route.get("products", []):
            name = prod.get("name", "")
            if name and name not in seen:
                seen.add(name)
                unique_prods.append(name)

    n_prod_cols = len(unique_prods)
    total_cols = 2 + n_prod_cols  # Маршрут + Адрес + продукты

    # Строка 1: заголовок
    title = f"Сборка по {dept_name} {date_str} {type_lbl}"
    if total_cols > 1:
        ws.write_merge(0, 0, 0, total_cols - 1, title, styles["title"])
    else:
        ws.write(0, 0, title, styles["title"])

    # Строка 2: заголовки столбцов
    ws.write(1, 0, "Маршрут", styles["header"])
    ws.write(1, 1, "Адрес", styles["header"])
    for ci, pname in enumerate(unique_prods):
        ws.write(1, 2 + ci, pname, styles["header"])

    # Данные
    routes_sorted = _sort_routes(routes, sort_asc)
    for ri, route in enumerate(routes_sorted):
        row = 2 + ri
        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")

        ws.write(row, 0, route_num_str, styles["header"])
        ws.write(row, 1, address, styles["header_wrap"])

        # Строим словарь продуктов маршрута для быстрого поиска
        prod_by_name: dict[str, dict] = {
            p.get("name", ""): p for p in route.get("products", [])
        }

        for ci, pname in enumerate(unique_prods):
            prod = prod_by_name.get(pname)
            if prod is not None:
                cell_val = _fmt_qty_with_pcs(prod)
            else:
                cell_val = ""
            ws.write(row, 2 + ci, cell_val, styles["num"])

    # Ширина столбцов
    _set_col_width(ws, 0, 14)
    _set_col_width(ws, 1, 42)
    for ci in range(n_prod_cols):
        _set_col_width(ws, 2 + ci, 20)


def _write_dept_rows(
    ws: xlwt.Worksheet,
    routes: list[dict],
    dept_name: str,
    date_str: str,
    type_lbl: str,
    styles: dict[str, xlwt.XFStyle],
    sort_asc: bool = False,
    prod_map: dict | None = None,
) -> None:
    """
    Формат 2 (rows): строчный формат — строка маршрута + строки продуктов.

    Структура:
      Строка 1: заголовок (объединённая)
      Строка 2: Маршрут | Адрес | Кол-во | Шт [| Грязные]
      Строка 3+: для каждого маршрута: строка маршрута + строки продуктов.

    При наличии продуктов с showInDirty добавляется колонка «Грязные» (×1,25).
    """
    prod_map = prod_map or {}
    has_dirty = any(
        prod_map.get(p.get("name", ""), {}).get("showInDirty")
        and data_store.is_subdept_chistchenka(prod_map.get(p.get("name", ""), {}).get("deptKey"))
        for r in routes for p in r.get("products", [])
    )
    n_cols = 5 if has_dirty else 4
    merge_cols = n_cols - 1

    # Строка 1: заголовок
    ws.write_merge(0, 0, 0, merge_cols, f"Сборка по {dept_name} {date_str} {type_lbl}",
                   styles["title"])

    # Строка 2: заголовки
    ws.write(1, 0, "Маршрут",    styles["header"])
    ws.write(1, 1, "Адрес",      styles["header"])
    ws.write(1, 2, "Кол-во",     styles["header"])
    ws.write(1, 3, "Шт",         styles["header"])
    if has_dirty:
        ws.write(1, 4, "Грязные", styles["header"])

    # Строка 3: единицы измерения
    ws.write(2, 0, "",           styles["header"])
    ws.write(2, 1, "",           styles["header"])
    ws.write(2, 2, "ед. изм.",   styles["header"])
    ws.write(2, 3, "шт",         styles["header"])
    if has_dirty:
        ws.write(2, 4, "",        styles["header"])

    routes_sorted = _sort_routes(routes, sort_asc)
    current_row = 3

    for route in routes_sorted:
        products = route.get("products", [])
        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")

        ws.write(current_row, 0, route_num_str, styles["header"])
        ws.write(current_row, 1, address,       styles["header_wrap"])
        ws.write(current_row, 2, "",            styles["cell"])
        ws.write(current_row, 3, "",            styles["cell"])
        if has_dirty:
            ws.write(current_row, 4, "",         styles["cell"])
        current_row += 1

        for prod in products:
            pname = prod.get("name", "")
            qty   = prod.get("displayQuantity", prod.get("quantity"))
            pcs   = prod.get("pcs")
            pcs_tail = prod.get("pcsTail")
            unit  = prod.get("unit", "")

            qty_str = f"{qty} {unit}".strip() if (qty is not None and unit) else (str(qty) if qty is not None else "")
            if pcs is not None:
                pcs_int = _fmt_pcs_int(pcs)
                if pcs_tail is not None and pcs_tail > 1e-9 and unit:
                    tail_txt = str(int(pcs_tail)) if abs(pcs_tail - round(pcs_tail)) < 1e-9 else f"{pcs_tail:.3f}".rstrip("0").rstrip(".")
                    pcs_str = f"{pcs_int} шт + {tail_txt} {unit}"
                else:
                    pcs_str = pcs_int or str(pcs)
            else:
                pcs_str = ""

            ws.write(current_row, 0, "",       styles["cell"])
            ws.write(current_row, 1, pname,    styles["cell"])
            ws.write(current_row, 2, qty_str,  styles["num"])
            ws.write(current_row, 3, pcs_str,  styles["num"])
            if has_dirty:
                dirty_val = ""
                ps = prod_map.get(pname, {})
                if ps.get("showInDirty") and data_store.is_subdept_chistchenka(ps.get("deptKey")):
                    try:
                        raw = float(prod.get("quantity", 0) or 0)
                        dirty = raw * 1.25
                        dirty_val = str(int(dirty)) if abs(dirty - round(dirty)) < 1e-9 else f"{dirty:.2f}"
                    except (ValueError, TypeError):
                        pass
                ws.write(current_row, 4, dirty_val, styles["num"])
            current_row += 1

    _set_col_width(ws, 0, 14)
    _set_col_width(ws, 1, 42)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 10)
    if has_dirty:
        _set_col_width(ws, 4, 12)


# ─────────────────────────── Устаревший формат (совместимость) ────────────

AVAILABLE_COLS: dict[str, str] = {
    "routeNumber":  "№ маршрута",
    "address":      "Адрес",
    "product":      "Продукт",
    "unit":         "Ед. изм.",
    "quantity":     "Количество",
    "pcs":          "Шт",
    "dirty":        "Грязные",
    "productQty":   "Продукт (кол-во)",
    "productsWide": "Продукт (колонка на каждый)",
    "nomenclature": "Номенклатура",
}

DEFAULT_COLS: list[dict] = [
    {"field": "routeNumber", "label": None, "merged": False},
    {"field": "address",     "label": None, "merged": False},
    {"field": "product",     "label": None, "merged": False},
    {"field": "unit",        "label": None, "merged": False},
    {"field": "quantity",    "label": None, "merged": False},
]

_COL_WIDTHS: dict[str, int] = {
    "routeNumber": 14, "address": 42, "product": 32,
    "unit": 12, "quantity": 14, "pcs": 8, "dirty": 12, "productQty": 32,
    "productsWide": 20, "nomenclature": 42,
}


def _get_col_label(col: dict, dept_key: str | None = None) -> str:
    if col.get("label"):
        return col["label"]
    if col.get("merged") and col.get("productName"):
        return col["productName"]
    lbl = AVAILABLE_COLS.get(col["field"], col.get("field", ""))
    if col.get("field") == "pcs" and dept_key and data_store.is_subdept_polufabricates(dept_key):
        return "кор."
    return lbl


DEFAULT_TEMPLATE_NAME = "Стандартный"

# Макс. строк данных на странице A4 при печати (шаблон 2). Заголовок ~3 строки, отступы ~2.5 см.
_ROWS_PER_A4_PAGE = 48


def is_template_2(tmpl: dict | None) -> bool:
    """Проверяет, включён ли «Шаблон 2» (Компактный) — для проверки влезания на A4."""
    if not tmpl:
        return False
    name = (tmpl.get("name") or "").strip().lower()
    tid = (tmpl.get("id") or "").lower()
    return (
        "шаблон 2" in name or "компактн" in name
        or tid == "template2"
    )


def dept_has_template_2(dept_key: str, templates: list[dict]) -> bool:
    """Проверяет, настроен ли для отдела/подотдела Шаблон 2 (разделение по продуктам только для него)."""
    tmpl = _get_template(dept_key, templates)
    return is_template_2(tmpl)


def _get_template(dept_key: str, templates: list[dict]) -> dict | None:
    """
    Возвращает шаблон для отдела или None. Только шаблоны с forDepartments=True.
    Если для отдела/подотдела не выбран шаблон — используется «Стандартный».
    """
    for tmpl in templates:
        if not tmpl.get("forDepartments", True):
            continue
        dept_keys = tmpl.get("deptKeys") or ([tmpl.get("deptKey")] if tmpl.get("deptKey") else [])
        if dept_key in dept_keys:
            return tmpl
    # Fallback: шаблон «Стандартный» для отделов без явной привязки
    for tmpl in templates:
        if (tmpl.get("name") or "").strip() == DEFAULT_TEMPLATE_NAME:
            return tmpl
    # Fallback: шаблон без привязки (deptKeys пустой)
    for tmpl in templates:
        if not tmpl.get("forDepartments", True):
            continue
        dept_keys = tmpl.get("deptKeys") or []
        if not dept_keys:
            return tmpl
    return templates[0] if templates else None


def get_template_name_for_dept(dept_key: str, templates: list[dict]) -> str:
    """Возвращает название шаблона для отдела/подотдела или «—» если шаблонов нет."""
    tmpl = _get_template(dept_key, templates)
    if tmpl:
        return (tmpl.get("name") or "").strip() or DEFAULT_TEMPLATE_NAME
    return "—"


def get_dept_preview_data(
    group: dict,
    prod_map: dict[str, dict],
    templates: list[dict],
    sort_asc: bool = True,
) -> tuple[list[str], list[str] | None, list[list[str]], set[tuple[int, int]]]:
    """
    Возвращает данные для предпросмотра отдела в формате шаблона.
    (headers, header_row2, rows, bold_cells) — bold_cells: set of (row_idx, col_idx) для жирного шт.
    """
    routes = [r for r in group["routes"] if not r.get("excluded", False)]
    _apply_pcs(routes, prod_map, group_dept_key=group.get("key"))
    prod_map = prod_map or {}
    dept_key = group.get("key", "")

    template_cols = _get_template_cols(dept_key, templates)
    template_cols = _inject_auto_columns(template_cols, routes, dept_key, prod_map)
    template_cols = _resolve_merged_cols(template_cols, routes)

    if any(c.get("field") == "nomenclature" for c in template_cols):
        # Номенклатура — только название продукта. После неё: ед. изм., количество, Шт (авто при округлении)
        has_pcs = any(p.get("pcs") is not None for r in routes for p in r.get("products", []))
        order = ("routeNumber", "nomenclature", "unit", "quantity", "pcs", "dirty")
        nm_cols = [c for c in template_cols if c.get("field") in order]
        if not nm_cols:
            nm_cols = [c for c in template_cols if c.get("field") == "nomenclature"]
        if nm_cols:
            by_field = {c.get("field"): c for c in nm_cols}
            template_cols = []
            for f in order:
                if f in by_field:
                    template_cols.append(by_field[f])
                elif f == "pcs" and has_pcs:
                    template_cols.append({"field": "pcs", "label": None, "merged": False})
                elif f in ("unit", "quantity"):
                    template_cols.append({"field": f, "label": None, "merged": False})

    headers = [
        "Номенклатура" if c.get("field") == "nomenclature" else _get_col_label(c, dept_key)
        for c in template_cols
    ]
    header_row2: list[str] | None = None
    has_product_qty = any(c.get("field") == "productQty" and c.get("productName") for c in template_cols)
    if has_product_qty:
        header_row2 = []
        for c in template_cols:
            if c.get("field") == "productQty" and c.get("productName"):
                unit = (prod_map.get(c["productName"], {}).get("unit") or "").strip() or "—"
                header_row2.append(unit)
            else:
                header_row2.append("")

    routes_sorted = _sort_routes(routes, sort_asc)
    rows: list[list[str]] = []
    bold_cells: set[tuple[int, int]] = set()
    has_nomenclature = any(c.get("field") == "nomenclature" for c in template_cols)
    has_row_per_product = any(c.get("field") in ("product", "unit", "quantity") for c in template_cols)
    tmpl = _get_template(dept_key, templates)

    def _cell_val(val) -> str:
        if val is None:
            return ""
        return str(val) if val != "" else ""

    for route in routes_sorted:
        products = route.get("products", [])
        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")

        if has_nomenclature:
            for pi in range(1 + len(products)):
                row_vals: list[str] = []
                is_address_row = pi == 0
                prod = products[pi - 1] if pi > 0 else None
                for ci, c in enumerate(template_cols):
                    f = c.get("field")
                    if f == "routeNumber":
                        row_vals.append(route_num_str if is_address_row else "")
                    elif f == "nomenclature":
                        if is_address_row:
                            row_vals.append(address)
                        elif prod:
                            # Только название продукта, без количества и ед. изм.
                            row_vals.append((prod.get("name") or "").strip())
                        else:
                            row_vals.append("")
                    elif f == "unit" and prod:
                        row_vals.append(_cell_val(prod.get("unit")))
                    elif f == "quantity" and prod:
                        qty = prod.get("displayQuantity", prod.get("quantity"))
                        row_vals.append(_cell_val(qty))
                    elif f == "pcs" and prod:
                        row_vals.append(_fmt_pcs_cell(prod, dept_key))
                        if _should_bold_pcs_cell(prod, tmpl, dept_key):
                            bold_cells.add((len(rows), ci))
                    elif f == "dirty" and prod:
                        ps = prod_map.get(prod.get("name", ""), {})
                        if ps.get("showInDirty") and data_store.is_subdept_chistchenka(ps.get("deptKey")):
                            try:
                                raw = float(prod.get("quantity", 0) or 0)
                                dirty = raw * 1.25
                                row_vals.append(str(int(dirty)) if abs(dirty - round(dirty)) < 1e-9 else f"{dirty:.2f}")
                            except (ValueError, TypeError):
                                row_vals.append("")
                        else:
                            row_vals.append("")
                    else:
                        row_vals.append("")
                rows.append(row_vals)
        else:
            n_prods = max(len(products), 1) if has_row_per_product else 1
            for pi in range(n_prods):
                prod = products[pi] if pi < len(products) else {}
                row_vals = []
                for ci, c in enumerate(template_cols):
                    f = c.get("field")
                    merged = c.get("merged", False)
                    if f == "routeNumber":
                        row_vals.append(route_num_str if pi == 0 else "")
                    elif f == "address":
                        row_vals.append(address if pi == 0 else "")
                    elif f == "product":
                        row_vals.append(_cell_val(prod.get("name")))
                    elif f == "unit":
                        row_vals.append(_cell_val(prod.get("unit")))
                    elif f == "quantity":
                        qty = prod.get("displayQuantity", prod.get("quantity"))
                        row_vals.append(_cell_val(qty))
                    elif f == "pcs":
                        row_vals.append(_fmt_pcs_cell(prod, dept_key))
                        if _should_bold_pcs_cell(prod, tmpl, dept_key):
                            bold_cells.add((len(rows), ci))
                    elif f == "dirty":
                        val = ""
                        ps = prod_map.get(prod.get("name", ""), {})
                        if ps.get("showInDirty") and data_store.is_subdept_chistchenka(ps.get("deptKey")):
                            try:
                                raw = float(prod.get("quantity", 0) or 0)
                                dirty = raw * 1.25
                                val = str(int(dirty)) if abs(dirty - round(dirty)) < 1e-9 else f"{dirty:.2f}"
                            except (ValueError, TypeError):
                                pass
                        row_vals.append(val)
                    elif f == "productQty" and merged:
                        target_name = c.get("productName", "")
                        target_prod = next((p for p in products if p.get("name", "") == target_name), None)
                        if target_prod and (not has_row_per_product or prod.get("name", "") == target_name):
                            row_vals.append(_fmt_qty_with_pcs(target_prod, dept_key))
                            if _should_bold_product_qty_cell(target_prod):
                                bold_cells.add((len(rows), ci))
                        else:
                            row_vals.append("")
                    else:
                        row_vals.append("")
                rows.append(row_vals)

    return headers, header_row2, rows, bold_cells


def _get_template_cols(dept_key: str, templates: list[dict]) -> list[dict]:
    """Возвращает список столбцов шаблона для отдела. При пустом columns — из grid."""
    tmpl = _get_template(dept_key, templates)
    if tmpl:
        cols = data_store.get_template_columns(tmpl)
        if cols:
            if isinstance(cols[0], str):
                return [{"field": c, "label": None, "merged": False} for c in cols]
            return cols
    return DEFAULT_COLS[:]


def _resolve_merged_cols(
    template_cols: list[dict],
    routes: list[dict],
) -> list[dict]:
    """
    Автоопределение объединённых столбцов (productQty).
    Без дубликатов: один столбец на продукт.
    Важно: productQty с productName добавляются только для продуктов из routes
    текущей группы — иначе при общем шаблоне (булки/пирожки и др.) в файл
    попадали бы колонки чужих подотделов.
    """
    unique_products: list[str] = []
    seen: set[str] = set()
    for route in routes:
        for prod in route.get("products", []):
            name = prod.get("name", "")
            if name and name not in seen:
                seen.add(name)
                unique_products.append(name)

    result: list[dict] = []
    product_qty_added: set[str] = set()

    def _add_product_qty(pname: str) -> None:
        if pname and pname not in product_qty_added:
            product_qty_added.add(pname)
            result.append({
                "field": "productQty",
                "label": pname,
                "merged": True,
                "productName": pname,
            })

    for col in template_cols:
        if col["field"] == "productsWide":
            for pname in unique_products:
                _add_product_qty(pname)
        elif col["field"] == "productQty":
            if col.get("productName"):
                # Добавляем только продукты текущей группы (из routes), не из других подотделов
                if col["productName"] in seen:
                    _add_product_qty(col["productName"])
            elif len(unique_products) == 1:
                _add_product_qty(unique_products[0])
            else:
                for pname in unique_products:
                    _add_product_qty(pname)
        else:
            result.append(col)
    return result


def _should_bold_pcs_cell(prod: dict | None, tmpl: dict | None, dept_key: str) -> bool:
    """
    Шт жирным: для шаблона 2 при showPcs (отдельный столбец), или для полуфабрикатов с хвостиком (pcsTail).
    При 0 шт + хвостик не жирно (шт не отображается).
    """
    if not prod:
        return False
    has_pcs = prod.get("pcs") is not None
    has_pcs_tail = prod.get("pcsTail") is not None and prod.get("pcsTail", 0) > 1e-9
    if has_pcs_tail and int(prod.get("pcs") or 0) > 0:
        return True
    if is_template_2(tmpl) and has_pcs:
        return True
    return False


def _should_bold_product_qty_cell(prod: dict | None) -> bool:
    """
    Для объединённой ячейки «количество / шт + хвостик»: жирным только когда отображаем
    «Шт + хвостик» (pcs > 0). При 0 шт + хвостик показываем только количество — не жирно.
    """
    if not prod:
        return False
    pcs = prod.get("pcs")
    tail = prod.get("pcsTail")
    if tail is None or tail <= 1e-9:
        return False
    return int(pcs or 0) > 0


def _write_dept_sheet_nomenclature(
    ws: xlwt.Worksheet,
    routes: list[dict],
    dept_name: str,
    date_str: str,
    type_lbl: str,
    template_cols: list[dict],
    styles: dict[str, xlwt.XFStyle],
    sort_asc: bool = False,
    prod_map: dict | None = None,
    split_for_a4: bool = False,
    tmpl: dict | None = None,
    dept_key: str = "",
) -> None:
    """
    Запись по шаблону с колонкой «Номенклатура»: заголовок «Номенклатура»,
    в первой строке блока — адрес, в следующих — продукты отдела.
    № маршрута только в строке с адресом.
    При split_for_a4: дублируем № маршрута и адрес в каждой строке, добавляем разрывы страниц.
    """
    routes_sorted = _sort_routes(routes, sort_asc)
    n_cols = len(template_cols)
    title = f"Сборка по {dept_name} {date_str} {type_lbl}"
    if n_cols > 1:
        ws.write_merge(0, 0, 0, n_cols - 1, title, styles["title"])
    else:
        ws.write(0, 0, title, styles["title"])

    for ci, col_def in enumerate(template_cols):
        lbl = "Номенклатура" if col_def.get("field") == "nomenclature" else _get_col_label(col_def, dept_key)
        ws.write(1, ci, lbl, styles["header"])

    current_row = 2
    page_breaks: list[int] = []
    rows_on_current_page = 0

    for route in routes_sorted:
        products = route.get("products", [])
        block_height = 1 + len(products)
        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")

        # Разрыв только перед блоком маршрута: новая страница — со строки с номером маршрута
        if split_for_a4 and rows_on_current_page > 0 and rows_on_current_page + block_height > _ROWS_PER_A4_PAGE:
            page_breaks.append(current_row)
            # Дублируем заголовок на каждой новой странице
            if n_cols > 1:
                ws.write_merge(current_row, current_row, 0, n_cols - 1, title, styles["title"])
            else:
                ws.write(current_row, 0, title, styles["title"])
            for ci, col_def in enumerate(template_cols):
                lbl = "Номенклатура" if col_def.get("field") == "nomenclature" else _get_col_label(col_def, dept_key)
                ws.write(current_row + 1, ci, lbl, styles["header"])
            current_row += 2
            rows_on_current_page = 0

        for pi in range(1 + len(products)):
            row = current_row + pi
            is_address_row = pi == 0
            prod = products[pi - 1] if pi > 0 else None

            for ci, col_def in enumerate(template_cols):
                field = col_def["field"]
                if field == "routeNumber":
                    val = route_num_str if (is_address_row or split_for_a4) else ""
                    ws.write(row, ci, val, styles["header"])
                elif field == "address":
                    # При split_for_a4 — адрес в каждой строке (копирование на страницу)
                    val = address if (is_address_row or split_for_a4) else ""
                    ws.write(row, ci, val, styles["header_wrap"])
                elif field == "nomenclature":
                    if is_address_row:
                        ws.write(row, ci, address, styles["header_wrap"])
                    elif prod is not None:
                        ws.write(row, ci, (prod.get("name") or "").strip(), styles["cell"])
                    else:
                        ws.write(row, ci, "", styles["cell"])
                elif field == "productQty":
                    ws.write(row, ci, "", styles["cell"])
                elif field == "product":
                    ws.write(row, ci, "", styles["cell"])
                elif field == "unit":
                    if prod is not None:
                        ws.write(row, ci, (prod.get("unit") or "").strip(), styles["cell"])
                    else:
                        ws.write(row, ci, "", styles["cell"])
                elif field == "quantity":
                    if prod is not None:
                        qty = prod.get("displayQuantity", prod.get("quantity"))
                        ws.write(row, ci, qty if qty is not None else "", styles["num"])
                    else:
                        ws.write(row, ci, "", styles["num"])
                elif field == "pcs":
                    if prod is not None:
                        style = styles["num_bold"] if _should_bold_pcs_cell(prod, tmpl, dept_key) else styles["num"]
                        ws.write(row, ci, _fmt_pcs_cell(prod, dept_key), style)
                    else:
                        ws.write(row, ci, "", styles["num"])
                elif field == "dirty" and prod is not None:
                    dirty_val = ""
                    ps = (prod_map or {}).get(prod.get("name", ""), {})
                    if ps.get("showInDirty") and data_store.is_subdept_chistchenka(ps.get("deptKey")):
                        try:
                            raw = float(prod.get("quantity", 0) or 0)
                            dirty = raw * 1.25
                            dirty_val = str(int(dirty)) if abs(dirty - round(dirty)) < 1e-9 else f"{dirty:.2f}"
                        except (ValueError, TypeError):
                            pass
                    ws.write(row, ci, dirty_val, styles["num"])
                else:
                    ws.write(row, ci, "", styles["cell"])

        current_row += block_height
        rows_on_current_page += block_height

    if split_for_a4 and page_breaks:
        ws.horz_page_breaks = [(r, 0, 255) for r in page_breaks]

    for ci, col_def in enumerate(template_cols):
        _set_col_width(ws, ci, _COL_WIDTHS.get(col_def["field"], 16))


def _write_dept_sheet(
    ws: xlwt.Worksheet,
    routes: list[dict],
    dept_name: str,
    date_str: str,
    type_lbl: str,
    template_cols: list[dict],
    styles: dict[str, xlwt.XFStyle],
    sort_asc: bool = False,
    prod_map: dict | None = None,
    split_for_a4: bool = False,
    tmpl: dict | None = None,
    dept_key: str = "",
) -> None:
    """Записывает данные на лист отдела по column-based шаблону."""
    template_cols = _resolve_merged_cols(template_cols, routes)
    prod_map = prod_map or {}
    if any(c.get("field") == "nomenclature" for c in template_cols):
        # Номенклатура — только название продукта. После неё: ед. изм., количество, Шт (при округлении)
        has_pcs = any(p.get("pcs") is not None for r in routes for p in r.get("products", []))
        order = ("routeNumber", "address", "nomenclature", "unit", "quantity", "pcs", "dirty")
        nm_cols = [c for c in template_cols if c.get("field") in order]
        if not nm_cols:
            nm_cols = [c for c in template_cols if c.get("field") == "nomenclature"]
        if nm_cols:
            by_field = {c.get("field"): c for c in nm_cols}
            template_cols = []
            for f in order:
                if f in by_field:
                    template_cols.append(by_field[f])
                elif f == "address" and split_for_a4:
                    template_cols.append({"field": "address", "label": "Адрес", "merged": False})
                elif f == "pcs" and has_pcs:
                    template_cols.append({"field": "pcs", "label": None, "merged": False})
                elif f in ("unit", "quantity"):
                    template_cols.append({"field": f, "label": None, "merged": False})
        _write_dept_sheet_nomenclature(
            ws, routes, dept_name, date_str, type_lbl, template_cols, styles, sort_asc, prod_map,
            split_for_a4=split_for_a4, tmpl=tmpl, dept_key=dept_key,
        )
        return
    n_cols = len(template_cols)

    title = f"Сборка по {dept_name} {date_str} {type_lbl}"
    if n_cols > 1:
        ws.write_merge(0, 0, 0, n_cols - 1, title, styles["title"])
    else:
        ws.write(0, 0, title, styles["title"])

    for ci, col_def in enumerate(template_cols):
        ws.write(1, ci, _get_col_label(col_def, dept_key), styles["header"])

    # Вторая строка заголовка: ед. изм. для столбцов «продукт (колонка на каждый)»
    has_product_qty = any(c.get("field") == "productQty" and c.get("productName") for c in template_cols)
    if has_product_qty:
        for ci, col_def in enumerate(template_cols):
            if col_def.get("field") == "productQty" and col_def.get("productName"):
                unit = (prod_map.get(col_def["productName"], {}).get("unit") or "").strip() or "—"
                ws.write(2, ci, unit, styles["header"])
            else:
                ws.write(2, ci, "", styles["header"])
        data_start_row = 3
    else:
        data_start_row = 2

    routes_sorted = _sort_routes(routes, sort_asc)
    current_row = data_start_row
    page_breaks: list[int] = []
    header_rows = data_start_row  # title + header (+ units)
    rows_on_current_page = 0

    # Только productQty (без product/unit/quantity) — одна строка на маршрут; иначе — строка на каждый продукт
    has_row_per_product = any(c.get("field") in ("product", "unit", "quantity") for c in template_cols)

    for route in routes_sorted:
        products = route.get("products", [])
        n_prods = max(len(products), 1) if has_row_per_product else 1
        block_height = n_prods
        route_num_str = str(route.get("routeNum", ""))
        address = route.get("address", "")

        # Разрыв только перед блоком маршрута: новая страница — со строки с номером маршрута
        if split_for_a4 and rows_on_current_page > 0 and rows_on_current_page + block_height > _ROWS_PER_A4_PAGE:
            page_breaks.append(current_row)
            # Дублируем заголовок на каждой новой странице
            if n_cols > 1:
                ws.write_merge(current_row, current_row, 0, n_cols - 1, title, styles["title"])
            else:
                ws.write(current_row, 0, title, styles["title"])
            for ci, col_def in enumerate(template_cols):
                ws.write(current_row + 1, ci, _get_col_label(col_def, dept_key), styles["header"])
            if has_product_qty:
                for ci, col_def in enumerate(template_cols):
                    if col_def.get("field") == "productQty" and col_def.get("productName"):
                        unit = (prod_map.get(col_def["productName"], {}).get("unit") or "").strip() or "—"
                        ws.write(current_row + 2, ci, unit, styles["header"])
                    else:
                        ws.write(current_row + 2, ci, "", styles["header"])
            current_row += header_rows
            rows_on_current_page = 0

        for pi in range(n_prods):
            prod = products[pi] if pi < len(products) else {}
            row = current_row + pi

            for ci, col_def in enumerate(template_cols):
                field = col_def["field"]
                merged = col_def.get("merged", False)

                if field == "routeNumber":
                    if split_for_a4:
                        ws.write(row, ci, route_num_str, styles["header"])
                    elif pi == 0:
                        if n_prods > 1:
                            ws.write_merge(row, row + n_prods - 1, ci, ci,
                                           route_num_str, styles["header"])
                        else:
                            ws.write(row, ci, route_num_str, styles["header"])

                elif field == "address":
                    if split_for_a4:
                        ws.write(row, ci, address, styles["header_wrap"])
                    elif pi == 0:
                        if n_prods > 1:
                            ws.write_merge(row, row + n_prods - 1, ci, ci,
                                           address, styles["header_wrap"])
                        else:
                            ws.write(row, ci, address, styles["header_wrap"])

                elif field == "product":
                    ws.write(row, ci, prod.get("name", ""), styles["cell"])

                elif field == "unit":
                    ws.write(row, ci, prod.get("unit", ""), styles["cell"])

                elif field == "quantity":
                    qty = prod.get("displayQuantity", prod.get("quantity"))
                    ws.write(row, ci, qty if qty is not None else "", styles["num"])

                elif field == "pcs":
                    style = styles["num_bold"] if _should_bold_pcs_cell(prod, tmpl, dept_key) else styles["num"]
                    ws.write(row, ci, _fmt_pcs_cell(prod, dept_key), style)

                elif field == "dirty":
                    dirty_val = ""
                    ps = prod_map.get(prod.get("name", ""), {})
                    if ps.get("showInDirty") and data_store.is_subdept_chistchenka(ps.get("deptKey")):
                        try:
                            raw = float(prod.get("quantity", 0) or 0)
                            dirty = raw * 1.25
                            dirty_val = str(int(dirty)) if abs(dirty - round(dirty)) < 1e-9 else f"{dirty:.2f}"
                        except (ValueError, TypeError):
                            pass
                    ws.write(row, ci, dirty_val, styles["num"])

                elif field == "productQty" and merged:
                    target_name = col_def.get("productName", "")
                    target_prod = next(
                        (p for p in products if p.get("name", "") == target_name),
                        None
                    )
                    # При строке на продукт — пишем только в столбец текущего продукта; иначе — все количества в одну строку
                    # Жирным: только для полуфабрикатов с хвостиком (в одной ячейке нельзя выделить только шт).
                    if target_prod is not None and (not has_row_per_product or prod.get("name", "") == target_name):
                        cell_text = _fmt_qty_with_pcs(target_prod, dept_key)
                        use_bold = _should_bold_product_qty_cell(target_prod)
                        ws.write(row, ci, cell_text, styles["num_bold"] if use_bold else styles["num"])
                    else:
                        ws.write(row, ci, "", styles["cell"])

                elif field in ("routeNumber", "address") and pi > 0:
                    pass  # объединённые ячейки — не перезаписывать

                else:
                    ws.write(row, ci, "", styles["cell"])

        current_row += block_height
        rows_on_current_page += block_height

    if split_for_a4 and page_breaks:
        ws.horz_page_breaks = [(r, 0, 255) for r in page_breaks]

    for ci, col_def in enumerate(template_cols):
        _set_col_width(ws, ci, _COL_WIDTHS.get(col_def["field"], 16))


# ─────────────────────────── Публичные функции генерации ──────────────────

def _inject_auto_columns(
    template_cols: list[dict],
    routes: list[dict],
    dept_key: str,
    prod_map: dict[str, dict],
) -> list[dict]:
    """
    Добавляет автоматические столбцы: Шт (при showPcs) и Грязные (для Чищенка при showInDirty).
    Удаляет pcs из шаблона — столбец добавляется только при необходимости.
    """
    has_pcs = any(
        p.get("pcs") is not None
        for r in routes
        for p in r.get("products", [])
    )
    has_dirty = (
        data_store.is_subdept_chistchenka(dept_key)
        and any(
            prod_map.get(p.get("name", ""), {}).get("showInDirty")
            for r in routes
            for p in r.get("products", [])
        )
    )
    result: list[dict] = []
    dirty_added = False
    for col in template_cols:
        if col.get("field") == "pcs":
            continue  # не из шаблона — добавляем автоматически
        result.append(col)
        if col.get("field") == "quantity":
            if has_pcs:
                pcs_label = _get_pcs_unit_label(dept_key)
                result.append({"field": "pcs", "label": pcs_label, "merged": False})
            if has_dirty:
                result.append({"field": "dirty", "label": "Грязные", "merged": False})
                dirty_added = True
    if has_dirty and not dirty_added:
        result.append({"field": "dirty", "label": "Грязные", "merged": False})
    return result


def _write_dept_by_format(
    ws: xlwt.Worksheet,
    routes: list[dict],
    dept_name: str,
    date_str: str,
    type_lbl: str,
    template_cols: list[dict],
    styles: dict[str, xlwt.XFStyle],
    sort_asc: bool = False,
    prod_map: dict | None = None,
    dept_key: str = "",
    template: dict | None = None,
    templates: list[dict] | None = None,
) -> None:
    """Записывает данные по шаблону (сетка). Авто-столбцы Шт и Грязные добавляются при необходимости."""
    prod_map = prod_map or {}
    template_cols = _inject_auto_columns(template_cols, routes, dept_key, prod_map)
    tmpl = template or (_get_template(dept_key, templates or []) if dept_key else None)
    # Разбиение на страницы A4 только для шаблона 2 и когда таблица не влезает
    has_nomenclature = any(c.get("field") == "nomenclature" for c in template_cols)
    has_row_per_product = any(c.get("field") in ("product", "unit", "quantity") for c in template_cols)
    if has_nomenclature:
        total_data_rows = sum(1 + len(r.get("products", [])) for r in routes)
    else:
        total_data_rows = sum(
            max(len(r.get("products", [])), 1) if has_row_per_product else 1
            for r in routes
        )
    split_for_a4 = is_template_2(tmpl) and total_data_rows > _ROWS_PER_A4_PAGE
    if split_for_a4:
        ws.paper_size_code = 9  # A4
    _write_dept_sheet(ws, routes, dept_name, date_str, type_lbl,
                     template_cols, styles, sort_asc, prod_map, split_for_a4=split_for_a4,
                     tmpl=tmpl, dept_key=dept_key)


def generate_single_dept_file(
    group: dict,
    file_type: str,
    save_path: str,
    prod_map: dict[str, dict],
    templates: list[dict],
    sort_asc: bool = False,
    replacements: list[dict] | None = None,
) -> str:
    """
    Создаёт файл для одного отдела/подотдела.

    Args:
        group: {"key", "name", "routes": [...]}
        file_type: "main" | "increase"
        save_path: полный путь к файлу .xls
        prod_map: dict {name: product_dict}
        templates: список шаблонов из data_store
        sort_asc: True — по возрастанию, False — по убыванию
    """
    routes = [r for r in group["routes"] if not r.get("excluded", False)]
    _apply_pcs(routes, prod_map, group_dept_key=group.get("key"))
    for r in routes:
        r["products"] = merge_replacement_pairs_for_display(
            r.get("products", []), replacements or []
        )

    template_cols = _get_template_cols(group["key"], templates)
    date_str = _format_date(_tomorrow())
    type_lbl = _type_label(file_type)

    wb = xlwt.Workbook(encoding="utf-8")
    styles = _get_styles()

    sheet_name = _safe_sheet_name(group["name"])
    ws = wb.add_sheet(sheet_name)
    _apply_page_margins(ws, for_labels=False)

    _write_dept_by_format(ws, routes, group["name"], date_str, type_lbl,
                          template_cols, styles, sort_asc, prod_map, group["key"],
                          templates=templates)

    _safe_save_workbook(wb, save_path)
    return save_path


def generate_dept_files(
    dept_groups: list[dict],
    file_type: str,
    save_dir: str,
    prod_map: dict[str, dict],
    templates: list[dict],
    sort_asc: bool = False,
    date_str: str | None = None,
    replacements: list[dict] | None = None,
) -> list[str]:
    """
    Создаёт файлы для всех отделов/подотделов.

    Args:
        dept_groups: список {"key", "name", "routes": [...]}
        file_type: "main" | "increase"
        save_dir: папка сохранения
        prod_map: dict {name: product_dict}
        templates: список шаблонов
        sort_asc: True — по возрастанию, False — по убыванию

    Returns:
        Список путей к созданным файлам.
    """
    date_str = date_str or get_routes_date_str()
    type_lbl = _type_label(file_type)

    created: list[str] = []
    styles = _get_styles()

    for group in dept_groups:
        parent = group.get("parent_dept_name") if group.get("is_subdept") else None
        save_path = get_dept_routes_path(
            save_dir, file_type, group["name"], date_str, parent_dept_name=parent
        )
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        routes = [r for r in group["routes"] if not r.get("excluded", False)]
        _apply_pcs(routes, prod_map, group_dept_key=group.get("key"))
        for r in routes:
            r["products"] = merge_replacement_pairs_for_display(
                r.get("products", []), replacements or []
            )

        template_cols = _get_template_cols(group["key"], templates)

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet(_safe_sheet_name(group["name"]))
        _apply_page_margins(ws, for_labels=False)
        _write_dept_by_format(ws, routes, group["name"], date_str, type_lbl,
                              template_cols, styles, sort_asc, prod_map, group["key"],
                              templates=templates)
        _safe_save_workbook(wb, save_path)
        created.append(save_path)

    return created


def generate_dept_files_by_products(
    dept_groups: list[dict],
    product_groups: dict[str, list[list[str]]],
    file_type: str,
    save_dir: str,
    prod_map: dict[str, dict],
    templates: list[dict],
    sort_asc: bool = False,
    date_str: str | None = None,
    replacements: list[dict] | None = None,
) -> list[str]:
    """
    Создаёт файлы по группам продуктов (режим «разделить по продуктам»).
    Разделение только для отделов/подотделов с Шаблоном 2. Остальные — один файл как обычно.
    product_groups: {dept_key: [[p1, p2], [p3], [p4, p5]]} — группы продуктов.
    """
    date_str = date_str or get_routes_date_str()
    type_lbl = _type_label(file_type)
    created: list[str] = []
    styles = _get_styles()

    for group in dept_groups:
        dept_key = group.get("key", "")
        parent = group.get("parent_dept_name") if group.get("is_subdept") else None
        base_routes = [r for r in group["routes"] if not r.get("excluded", False)]

        if not dept_has_template_2(dept_key, templates):
            # Шаблон не 2 — один файл на отдел (как в generate_dept_files)
            routes = copy.deepcopy(base_routes)
            _apply_pcs(routes, prod_map, group_dept_key=dept_key)
            for r in routes:
                r["products"] = merge_replacement_pairs_for_display(
                    r.get("products", []), replacements or []
                )
            save_path = get_dept_routes_path(
                save_dir, file_type, group["name"], date_str, parent_dept_name=parent
            )
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            template_cols = _get_template_cols(dept_key, templates)
            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet(_safe_sheet_name(group["name"]))
            _apply_page_margins(ws, for_labels=False)
            _write_dept_by_format(ws, routes, group["name"], date_str, type_lbl,
                                  template_cols, styles, sort_asc, prod_map, dept_key,
                                  templates=templates)
            _safe_save_workbook(wb, save_path)
            created.append(save_path)
            continue

        unique_products: list[str] = []
        seen: set[str] = set()
        for r in base_routes:
            for p in r.get("products", []):
                name = p.get("name", "")
                if name and name not in seen:
                    seen.add(name)
                    unique_products.append(name)

        groups_list = product_groups.get(dept_key)
        if not groups_list:
            groups_list = [[p] for p in unique_products]

        for product_names in groups_list:
            if not product_names:
                continue
            prod_set = set(product_names)
            routes_filtered = []
            for r in base_routes:
                dept_prods = [p for p in r.get("products", []) if p["name"] in prod_set]
                if not dept_prods:
                    continue
                routes_filtered.append({
                    "routeNum": r.get("routeNum", ""),
                    "address": r.get("address", ""),
                    "routeCategory": r.get("routeCategory") or "ШК",
                    "products": list(dept_prods),
                })

            if not routes_filtered:
                continue

            routes = copy.deepcopy(routes_filtered)
            _apply_pcs(routes, prod_map, group_dept_key=dept_key)
            for r in routes:
                r["products"] = merge_replacement_pairs_for_display(
                    r.get("products", []), replacements or []
                )

            save_path = get_dept_product_file_path(
                save_dir, file_type, group["name"], product_names, date_str,
                parent_dept_name=parent,
            )
            os.makedirs(os.path.dirname(save_path), exist_ok=True)

            template_cols = _get_template_cols(dept_key, templates)
            sheet_title = ", ".join(product_names) if len(product_names) <= 3 else f"{product_names[0]} и др."
            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet(_safe_sheet_name(sheet_title))
            _apply_page_margins(ws, for_labels=False)
            _write_dept_by_format(ws, routes, group["name"], date_str, type_lbl,
                                  template_cols, styles, sort_asc, prod_map, dept_key,
                                  templates=templates)
            _safe_save_workbook(wb, save_path)
            created.append(save_path)

    return created


def _aggregate_pcs_totals_by_product(
    routes: list[dict],
    prod_map: dict[str, dict],
) -> dict[tuple[str, str, str], float]:
    data = copy.deepcopy(routes or [])
    _apply_pcs(data, prod_map)
    totals: dict[tuple[str, str, str], float] = {}
    for route in data:
        if route.get("excluded"):
            continue
        for prod in route.get("products", []):
            name = prod.get("name") or ""
            if not name:
                continue
            settings = prod_map.get(name) or {}
            if not settings.get("showPcs"):
                continue
            dept_key = str(settings.get("deptKey") or "")
            unit = str(prod.get("unit") or settings.get("unit") or "")
            pcs = prod.get("pcs")
            if pcs is None:
                continue
            try:
                key = (dept_key, name, unit)
                totals[key] = totals.get(key, 0.0) + float(pcs)
            except (TypeError, ValueError):
                continue
    return totals


def _write_single_pcs_report(
    report_path: str,
    routes: list[dict],
    products_ref: list[dict],
    type_label: str,
    date_str: str,
) -> str:
    """Создаёт один отчёт по Шт для типа Основные или Увеличение."""
    prod_map = {p.get("name"): dict(p) for p in (products_ref or []) if p.get("name")}
    totals = _aggregate_pcs_totals_by_product(routes, prod_map)

    enabled_products: list[tuple[str, str, str]] = []
    for p in (products_ref or []):
        name = p.get("name")
        if not name or not p.get("showPcs"):
            continue
        enabled_products.append((str(p.get("deptKey") or ""), name, str(p.get("unit") or "")))

    all_keys = sorted(
        set(enabled_products) | set(totals.keys()),
        key=lambda x: (
            (data_store.get_department_display_name(x[0]) or "").lower(),
            x[1].lower(),
        ),
    )

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Отчет Шт")
    _apply_page_margins(ws, for_labels=False)
    styles = _get_styles()

    ws.write_merge(0, 0, 0, 4, f"Отчет по Шт {type_label} {date_str}", styles["title"])
    ws.write(1, 0, "Отдел / Подотдел", styles["header"])
    ws.write(1, 1, "Продукт", styles["header"])
    ws.write(1, 2, "количество в 1 шт", styles["header"])
    ws.write(1, 3, "Шт", styles["header"])
    ws.write(1, 4, "Итого", styles["header"])

    row = 2
    for dept_key, name, unit in all_keys:
        dept_name = data_store.get_department_display_name(dept_key) if dept_key else "Без отдела"
        m = totals.get((dept_key, name, unit), 0.0)
        settings = prod_map.get(name) or {}
        pcu = float(settings.get("pcsPerUnit", 1) or 1)
        if pcu <= 0:
            pcu = 1.0
        qty_in_one = f"{pcu} {unit}".strip() if unit else str(pcu)
        total_units = m * pcu

        ws.write(row, 0, dept_name, styles["cell"])
        ws.write(row, 1, name, styles["cell"])
        ws.write(row, 2, qty_in_one, styles["cell"])
        ws.write(row, 3, int(m) if abs(m - round(m)) < 1e-9 else round(m, 1), styles["num"])
        total_val = int(total_units) if abs(total_units - round(total_units)) < 1e-9 else round(total_units, 1)
        ws.write(row, 4, f"{total_val} {unit}".strip() if unit else str(total_val), styles["cell"])
        row += 1

    _set_col_width(ws, 0, 28)
    _set_col_width(ws, 1, 36)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 16)
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    wb.save(report_path)
    return report_path


def generate_pcs_compare_report(
    day_dir: str,
    main_routes: list[dict],
    increase_routes: list[dict],
    products_ref: list[dict],
    date_str: str | None = None,
) -> list[str]:
    """
    Создаёт отдельные отчёты по Шт для Основные и Увеличение.
    Файлы сохраняются в папках Основные и Увеличение соответственно.
    Возвращает список путей к созданным файлам.
    """
    date_str = date_str or get_routes_date_str()
    created: list[str] = []

    main_type_dir = os.path.join(day_dir, "Основные")
    main_path = os.path.join(main_type_dir, f"Отчет Шт Основные {date_str}.xls")
    _write_single_pcs_report(
        main_path,
        main_routes or [],
        products_ref or [],
        "Основные",
        date_str,
    )
    created.append(main_path)

    inc_type_dir = os.path.join(day_dir, "Увеличение")
    inc_path = os.path.join(inc_type_dir, f"Отчет Шт Увеличение {date_str}.xls")
    _write_single_pcs_report(
        inc_path,
        increase_routes or [],
        products_ref or [],
        "Увеличение",
        date_str,
    )
    created.append(inc_path)

    return created
